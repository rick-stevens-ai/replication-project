#!/usr/bin/env python3
"""Build the independent Table S2 systems JSON directly from xlsx and cross-check
against the replication's parsed version."""
import json, re
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parents[3].parent / "work" / "SupplementaryTables.xlsx"
OUT = Path(__file__).resolve().parents[1] / "data"
REPL = Path(__file__).resolve().parents[2] / "paper_S2_systems.json"

wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb["Table S2"]
rows = list(ws.iter_rows(values_only=True))

entries = []
for r in rows[2:]:  # skip title + header
    if not r or r[0] is None: continue
    pd, clone, source, contig, c1, c2, c3, start, stop, rc_s, rc_e = (list(r) + [None]*11)[:11]
    proteins = [c for c in [c1,c2,c3] if c and str(c).strip() not in ("", "No ID", "None", "-")]
    entries.append({
        "pd": pd,
        "clone": clone,
        "source": source,
        "contig": contig,
        "proteins": proteins,
        "start": start,
        "stop": stop,
    })

print(f"Independent S2 systems parsed: {len(entries)}")
total_prot = sum(len(e["proteins"]) for e in entries)
print(f"Independent S2 total protein components: {total_prot}")
unique_contigs = len(set(e["contig"] for e in entries))
unique_sources = len(set(e["source"] for e in entries))
print(f"unique contigs: {unique_contigs}")
print(f"unique source strains: {unique_sources}")

with open(OUT / "indep_s2_systems.json", "w") as fh:
    json.dump(entries, fh, indent=2, ensure_ascii=False)

# Cross-check vs replication
with open(REPL) as fh:
    repl = json.load(fh)

def norm_pd(s):
    return s.replace("λ","\u03bb").replace("\u03bb","λ")

by_pd = {e["pd"]: e for e in entries}
by_pd_repl = {r["pd"]: r for r in repl}

print("\nCross-check indep vs replication:")
mism = 0
for pd in sorted(by_pd):
    ind = by_pd[pd]
    r = by_pd_repl.get(pd)
    if not r:
        print(f"  {pd}: MISSING in replication")
        mism += 1; continue
    same_contig = ind["contig"] == r["contig_acc"]
    ind_prot = set(ind["proteins"])
    r_prot = set(r["cds"])
    same_prot = ind_prot == r_prot
    same_src = ind["source"] == r["source"]
    same_start = ind["start"] == r["coords"]
    ok = same_contig and same_prot and same_src and same_start
    if not ok:
        mism += 1
    print(f"  {pd}: src={'OK' if same_src else 'MISM'} contig={'OK' if same_contig else 'MISM'} prot={'OK' if same_prot else 'MISM'} start={'OK' if same_start else 'MISM'}")

print(f"\nmismatches: {mism}/{len(by_pd)}")
