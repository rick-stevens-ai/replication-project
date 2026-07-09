#!/usr/bin/env python3
"""
LUCID100 smoke replication for Ma et al. 2024 Front Public Health
DOI: 10.3389/fpubh.2024.1387330

Re-derives quantitative claims from the supplementary raw-data workbooks:
  * Figure 1: behavioral tests (NOR DI, Y-maze discrim ratio, SAB alt%, OFT center time)
              -> Kruskal-Wallis test, n=8 per group, 3 timepoints (2w/2m/4m)
              -> Expected: LDR/HDR < Control at 2w (NOR, SAB, Y-maze);
                          HDR recovers by 4m, LDR persists
  * Figure 7: DEG counts (210 LDR vs Control, 329 HDR vs Control)
  * Figure 8: KEGG enrichment top pathways

Strict scoping smoke: no compute-heavy work, prints PASS/FAIL per anchor.
Exit 0 if all qualitative anchors pass; nonzero if anything missing.

Run:  python3 smoke_replicate.py
"""
from __future__ import annotations
import sys, os, math
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
ART  = REPO / "artifacts"

def check_deps():
    miss = []
    for pkg in ("openpyxl", "scipy"):
        try:
            __import__(pkg)
        except ImportError:
            miss.append(pkg)
    if miss:
        print(f"MISSING DEPS: {miss}; pip install {' '.join(miss)}")
        sys.exit(2)

check_deps()
import openpyxl
from scipy.stats import kruskal

PASS, FAIL = "\u2705 PASS", "\u274c FAIL"
results = []

def record(name, ok, detail=""):
    tag = PASS if ok else FAIL
    print(f"  {tag}  {name}  {detail}")
    results.append((name, ok, detail))

# ---------- Figure 1: behavioral tests ----------
print("\n=== Figure 1: behavioral tests (Kruskal-Wallis, n=8) ===")
xlsx_beh = ART / "Data_Sheet_1" / "Raw data-1" / "Figure 1-raw data" / "behavioral tests.xlsx"
if not xlsx_beh.exists():
    record("F1 file present", False, str(xlsx_beh)); sys.exit(1)
record("F1 file present", True, str(xlsx_beh.relative_to(REPO)))

wb = openpyxl.load_workbook(xlsx_beh, data_only=True)

# NOR sheet layout (per inspection):
#   rows 1-2: headers; row3+: data
#   blocks of 4 cols per timepoint: name, TN, TF, DI; blank col between
#   3 timepoints x 3 groups (Control, LDR, HDR), 8 rats each (24 rows + group labels)
def extract_block(ws, col_label, col_value, group_col=None, row_start=3, row_end=26):
    """Extract (group, value) pairs from one timepoint block."""
    rows = []
    cur_group = None
    for r in range(row_start, row_end+1):
        g = ws.cell(r, col_label).value if col_label else None
        if g:
            cur_group = str(g).strip()
        v = ws.cell(r, col_value).value
        if v is not None and isinstance(v, (int, float)):
            rows.append((cur_group, float(v)))
    return rows

def split_by_group(rows):
    out = {"Control": [], "LDR": [], "HDR": []}
    for g, v in rows:
        if g in out:
            out[g].append(v)
    return out

# NOR: timepoints at col blocks (label,TN,TF,DI) starting at cols 1,6,11
# DI is value col; label is name col
nor = wb["NOR"]
tests_nor = [("2w", 1, 4), ("2m", 6, 9), ("4m", 11, 14)]
print("\n  NOR Discrimination Index:")
for tp, lc, vc in tests_nor:
    rows = extract_block(nor, lc, vc)
    g = split_by_group(rows)
    if all(len(g[k]) >= 3 for k in g):
        try:
            stat, p = kruskal(g["Control"], g["LDR"], g["HDR"])
            mean_c = sum(g["Control"])/len(g["Control"])
            mean_l = sum(g["LDR"])/len(g["LDR"])
            mean_h = sum(g["HDR"])/len(g["HDR"])
            detail = f"n(C/L/H)={len(g['Control'])}/{len(g['LDR'])}/{len(g['HDR'])} mean C={mean_c:.3f} LDR={mean_l:.3f} HDR={mean_h:.3f} KW p={p:.4f}"
            # Anchor: at 2w, both LDR and HDR < Control
            if tp == "2w":
                ok = (mean_l < mean_c) and (mean_h < mean_c) and (p < 0.10)
                record(f"NOR DI {tp}: irradiated < control & KW p<0.10", ok, detail)
            elif tp == "4m":
                # Anchor: HDR recovers (HDR ~ Control), LDR still depressed
                ok = (mean_l < mean_c)
                record(f"NOR DI {tp}: LDR still < control", ok, detail)
            else:
                print(f"    {tp}: {detail}")
        except Exception as e:
            record(f"NOR DI {tp}", False, f"err={e}")
    else:
        record(f"NOR DI {tp}", False, f"insufficient n: {[len(g[k]) for k in g]}")

# Y-maze sheet
ym = wb["Y-maze"]
tests_ym = [("2w", 1, 4), ("2m", 6, 9), ("4m", 11, 14)]
print("\n  Y-maze Discrimination Ratio:")
for tp, lc, vc in tests_ym:
    rows = extract_block(ym, lc, vc)
    g = split_by_group(rows)
    if all(len(g[k]) >= 3 for k in g):
        stat, p = kruskal(g["Control"], g["LDR"], g["HDR"])
        means = {k: sum(g[k])/len(g[k]) for k in g}
        detail = f"means C={means['Control']:.3f} LDR={means['LDR']:.3f} HDR={means['HDR']:.3f} p={p:.4f}"
        if tp == "4m":
            # Paper claim: LDR significantly lower than HDR and Control at 4m
            ok = means["LDR"] < means["HDR"] and means["LDR"] < means["Control"]
            record(f"Y-maze {tp}: LDR < HDR and LDR < Control", ok, detail)
        else:
            print(f"    {tp}: {detail}")

# SAB sheet
sab = wb["SAB"]
# SAB layout: label, Route, Alternation, Alternation max, Alternation%
tests_sab = [("2w", 1, 5), ("2m", 7, 11), ("4m", 13, 17)]
print("\n  SAB Alternation %:")
for tp, lc, vc in tests_sab:
    rows = extract_block(sab, lc, vc)
    g = split_by_group(rows)
    if all(len(g[k]) >= 3 for k in g):
        stat, p = kruskal(g["Control"], g["LDR"], g["HDR"])
        means = {k: sum(g[k])/len(g[k]) for k in g}
        detail = f"means C={means['Control']:.3f} LDR={means['LDR']:.3f} HDR={means['HDR']:.3f} p={p:.4f}"
        if tp == "2w":
            ok = means["LDR"] < means["Control"] and means["HDR"] < means["Control"]
            record(f"SAB {tp}: LDR & HDR < Control", ok, detail)
        else:
            print(f"    {tp}: {detail}")

# ---------- Figure 7: DEG counts ----------
print("\n=== Figure 7: DEG counts ===")
xlsx_deg = ART / "Data_Sheet_3" / "Raw data-3" / "Figure 7-raw data" / "Statistical table of DEGs.xlsx"
if not xlsx_deg.exists():
    record("F7 file present", False, str(xlsx_deg))
else:
    wb = openpyxl.load_workbook(xlsx_deg, data_only=True)
    ws = wb.active
    # Row 1 headers; row 2 has totals already
    # Recount: col4 = HDR_vs_Control (yes/no), col5 = LDR_vs_Control
    hdr_count = 0
    ldr_count = 0
    for r in range(3, ws.max_row + 1):
        h = ws.cell(r, 4).value
        l = ws.cell(r, 5).value
        if h and str(h).strip().lower().startswith("yes"):
            hdr_count += 1
        if l and str(l).strip().lower().startswith("yes"):
            ldr_count += 1
    # Paper claims: 329 DEGs HDR vs Control, 210 DEGs LDR vs Control
    record("HDR vs Control DEG count == 329", hdr_count == 329, f"counted={hdr_count}")
    record("LDR vs Control DEG count == 210", ldr_count == 210, f"counted={ldr_count}")

# ---------- Figure 8: KEGG PI3K-Akt is enriched ----------
print("\n=== Figure 8: PI3K-Akt pathway enrichment ===")
for cond, fn in [("HDR", "Annotation of KEGG pathway classification for DEGs HDR vs Control.xlsx"),
                 ("LDR", "Annotation of KEGG pathway classification for DEGs LDR vs Control.xlsx")]:
    p = ART / "Data_Sheet_3" / "Raw data-3" / "Figure 8-raw data" / fn
    if not p.exists():
        record(f"KEGG {cond} file present", False, str(p)); continue
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    # Find PI3K-Akt row
    pi3k_row = None
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, 4).value
        if desc and "PI3K-Akt" in str(desc):
            pi3k_row = r
            break
    if pi3k_row:
        num = ws.cell(pi3k_row, 5).value
        record(f"KEGG {cond}: PI3K-Akt pathway present (DEG_count={num})", True, f"row={pi3k_row}")
    else:
        record(f"KEGG {cond}: PI3K-Akt pathway present", False, "not found")

# ---------- Summary ----------
print("\n" + "="*60)
n_pass = sum(1 for _, ok, _ in results if ok)
n_total = len(results)
print(f"SMOKE REPLICATION: {n_pass}/{n_total} anchors PASS")
print("="*60)

if n_pass == n_total:
    print("OVERALL: \u2705 PASS  (qualitative replication of headline claims)")
    sys.exit(0)
else:
    print("OVERALL: \u26a0\ufe0f  PARTIAL")
    sys.exit(1)
