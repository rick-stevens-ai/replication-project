#!/usr/bin/env python3
"""Build distribution_summary.tsv combining BLAST, HMM, and genomic context data."""
import os, json, csv

BASEDIR = os.path.expanduser("~/Dropbox/REPLICATE-PROJECT/36123438-Anti-phage-defense-Ecoli")

# Load BLAST summary
blast_summary_file = os.path.join(BASEDIR, "analysis/blast_summary.json")
with open(blast_summary_file) as f:
    blast_summary = json.load(f)

# Load genomic context
context_file = os.path.join(BASEDIR, "analysis/genomic_context/context_summary.json")
with open(context_file) as f:
    context = json.load(f)

# Load defense island summary
defense_file = os.path.join(BASEDIR, "analysis/defense_island_summary.json")
with open(defense_file) as f:
    defense = json.load(f)

# Map system names (context uses PD-L instead of PD-λ)
def normalize_name(name):
    return name.replace("PD-L-", "PD-λ-")

# Parse HMM hits for the PD systems themselves (not context)
hmm_hits = {}
with open(os.path.join(BASEDIR, "analysis/PD_vs_DefenseDomains.tbl")) as f:
    for line in f:
        if line.startswith("#"):
            continue
        parts = line.split()
        if len(parts) < 18:
            continue
        domain = parts[0]
        pfam = parts[1] if parts[1] != "-" else ""
        protein = parts[2]
        evalue = float(parts[4])
        hmm_hits.setdefault(protein, []).append(f"{domain}({pfam})")

# System -> protein map
system_proteins = {}
with open(os.path.join(BASEDIR, "data/system_protein_map.tsv")) as f:
    next(f)
    for line in f:
        parts = line.strip().split("\t")
        if len(parts) == 2:
            system_proteins.setdefault(parts[0], []).append(parts[1])

# Read SorekZhang BLAST results
sorek_hits = {}
with open(os.path.join(BASEDIR, "analysis/PD_vs_SorekZhang.blastp")) as f:
    for line in f:
        parts = line.strip().split("\t")
        sorek_hits.setdefault(parts[0], []).append(parts[1])

# Read Table S1 system descriptions and Table S4 Gao et al. matches
import openpyxl
wb = openpyxl.load_workbook(os.path.join(BASEDIR, "data/SupplementaryTables.xlsx"), data_only=True)

# Table S1 - system type
system_type = {}
ws = wb["Table S1"]
for row in ws.iter_rows(values_only=True, min_row=3):
    name = str(row[0]) if row[0] else ""
    if name.startswith("PD-"):
        system_type[name] = str(row[2]) if row[2] else "?"

# Table S4 - Gao et al. matches
gao_matches = {}
ws = wb["Table S4"]
for row in ws.iter_rows(values_only=True, min_row=3):
    name = str(row[0]) if row[0] else ""
    if name.startswith("PD-"):
        base = name.rstrip("ABC")
        if base not in gao_matches:
            gao_matches[base] = []
        cluster = str(row[1]) if row[1] else "NA"
        pident = str(row[2]) if row[2] else "NA"
        if cluster != "NA":
            gao_matches[base].append(f"{cluster}({pident}%)")

# All 21 systems
all_systems = sorted(system_proteins.keys(), 
    key=lambda s: (s.split("-")[1].replace("λ", "L"), int(s.split("-")[2]) if s.split("-")[2].isdigit() else 0))

# Build output
outpath = os.path.join(BASEDIR, "analysis/distribution_summary.tsv")
with open(outpath, "w", newline="") as f:
    writer = csv.writer(f, delimiter="\t")
    writer.writerow([
        "System", "Phage_target", "System_type", "Num_genes",
        "Ecoli_homologs", "Total_homologs", "Num_organisms",
        "Phage_CDS_in_context", "Context_CDS_total", "Prophage_classification",
        "Known_defense_BLAST_hit", "DefenseDomain_HMM",
        "DISign_positive_domains", "Gao_et_al_cluster",
        "Defense_neighbors_BLAST", "Defense_neighbors_HMM",
        "Defense_island_context"
    ])
    
    for sys in all_systems:
        phage_target = sys.split("-")[1]
        sys_desc = system_type.get(sys, "?")
        num_genes = len(system_proteins[sys])
        
        # BLAST distribution
        blast_key = sys
        if blast_key in blast_summary:
            ecoli_h = blast_summary[blast_key]["ecoli_filtered"]
            all_h = blast_summary[blast_key]["all_filtered"]
            orgs = blast_summary[blast_key]["organisms"]
        else:
            ecoli_h = "pending"
            all_h = "pending"
            orgs = "pending"
        
        # Context
        ctx_key = sys.replace("λ", "L")
        if ctx_key in context:
            phage_cds = context[ctx_key]["phage_cds_count"]
            total_cds = context[ctx_key]["cds_count"]
            if phage_cds >= 5:
                prophage_class = "IN_PROPHAGE"
            elif phage_cds >= 1:
                prophage_class = "NEAR_MGE"
            else:
                prophage_class = "NO_PHAGE"
        else:
            phage_cds = "?"
            total_cds = "?"
            prophage_class = "?"
        
        # Known defense BLAST
        has_sorek = "NO"
        for prot in system_proteins[sys]:
            if prot in sorek_hits:
                has_sorek = "YES: " + ",".join(set(sorek_hits[prot]))
        
        # HMM domains
        domains = set()
        for prot in system_proteins[sys]:
            if prot in hmm_hits:
                domains.update(hmm_hits[prot])
        hmm_str = "; ".join(sorted(domains)) if domains else "NONE"
        
        # DISign - count positive domains from HMM hits
        disign_count = 0
        disign = {}
        with open(os.path.join(BASEDIR, "data/phagedefense/DISign.txt")) as df:
            for line in df:
                parts = line.strip().split("\t")
                if len(parts) >= 2:
                    disign[parts[0]] = parts[1]
        for d in domains:
            # Extract PF accession from domain(PFxxxxx.y)
            if "PF" in d:
                import re
                pfam_match = re.search(r'(PF\d+)', d)
                if pfam_match:
                    pfam = pfam_match.group(1)
                    if pfam in disign and disign[pfam] == "positive":
                        disign_count += 1
        
        # Gao et al.
        gao = "; ".join(gao_matches.get(sys, [])) if sys in gao_matches else "NA"
        
        # Defense neighbors
        def_key = ctx_key
        if def_key in defense:
            def_blast = defense[def_key]["blast_defense_neighbors"]
            def_hmm = defense[def_key]["hmm_defense_neighbors"]
            def_disign = defense[def_key]["disign_positive_neighbors"]
        else:
            def_blast = "?"
            def_hmm = "?"
            def_disign = "?"
        
        # Defense island context classification
        if def_blast > 0 or def_disign >= 3:
            di_context = "DEFENSE_ISLAND"
        elif def_disign >= 1:
            di_context = "PARTIAL_DI"
        else:
            di_context = "NON_DI"
        
        writer.writerow([
            sys, phage_target, sys_desc, num_genes,
            ecoli_h, all_h, orgs,
            phage_cds, total_cds, prophage_class,
            has_sorek, hmm_str,
            disign_count, gao,
            def_blast, def_hmm,
            di_context
        ])

print(f"Wrote {outpath}")

# Also print summary stats
print(f"\n=== Key Statistics ===")
print(f"Total systems: {len(all_systems)}")

# Distribution breadth
available = [s for s in all_systems if s in blast_summary]
ecoli_counts = [blast_summary[s]["ecoli_filtered"] for s in available]
print(f"\nDistribution (E. coli homologs, {len(available)}/21 systems):")
print(f"  Median: {sorted(ecoli_counts)[len(ecoli_counts)//2]}")
print(f"  Range: {min(ecoli_counts)} - {max(ecoli_counts)}")
print(f"  Broad (≥50 E. coli genomes): {sum(1 for c in ecoli_counts if c >= 50)}")
print(f"  Narrow (<20 E. coli genomes): {sum(1 for c in ecoli_counts if c < 20)}")
