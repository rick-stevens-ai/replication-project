#!/usr/bin/env python3
"""
RE-PASS replication script for Jiang et al. 2017 (PMID 28589945).

Goal: lift COVERAGE on the prior partial verdict by attacking claims pass-1 marked
NOT_TESTED or PARTIAL. Single script with incremental outputs in results/repass/.

Compute: free (local CPU + NCBI E-utilities, no paid APIs).
Provenance: every number traces back to an exact file/cell/accession.
"""

import os, sys, json, time, csv, gzip, subprocess, urllib.parse, urllib.request, hashlib
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]   # 28589945-ARG-dissemination/
RESULTS = ROOT / 'results' / 'repass'
RESULTS.mkdir(parents=True, exist_ok=True)
DATA_V2 = ROOT / 'data_v2'
PAPER = ROOT / 'paper'

def write_json(name, obj):
    p = RESULTS / name
    p.write_text(json.dumps(obj, indent=2, default=str))
    print(f'[wrote] {p.relative_to(ROOT)}')
    return p

def log(msg):
    print(f'[{time.strftime("%H:%M:%S")}] {msg}', flush=True)

# ----------------------------------------------------------------------
# CLAIM C2: Parser provenance — re-derive entry counts from Supp Data 1
# ----------------------------------------------------------------------
def claim_C2_parser_provenance():
    log('C2: re-parsing Supplementary Data 1')
    import openpyxl
    wb = openpyxl.load_workbook(PAPER / 'supp_data1.xlsx', data_only=True)
    ws = wb['Supplementary Data 1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    accs_col0 = [r[0] for r in rows if r[0]]
    uniq_col0 = sorted(set(accs_col0))

    # Strip the "99% identical to..." noise: take just first token
    def normalize_acc(a):
        if not a: return None
        a = str(a).strip().split()[0]
        return a

    norm = [normalize_acc(a) for a in accs_col0]
    norm_uniq = sorted(set(x for x in norm if x))

    # Identity range
    idents = [r[13] for r in rows if isinstance(r[13], (int, float))]

    # Y marks
    y_more_sim = [r for r in rows if r[17] and str(r[17]).strip().upper().startswith('Y')]

    # self-protecting / in cluster
    sp_or_cluster = [
        r for r in rows
        if r[6] and str(r[6]).strip().lower() not in ('', 'no', '-')
    ]

    result = {
        'source_file': 'paper/supp_data1.xlsx',
        'sheet': 'Supplementary Data 1',
        'data_rows': len(rows),
        'raw_accession_col_unique': len(uniq_col0),
        'normalized_accession_unique': len(norm_uniq),
        'identity_min': min(idents),
        'identity_max': max(idents),
        'identity_n': len(idents),
        'y_more_similar_to_actino': len(y_more_sim),
        'self_protecting_or_in_cluster': len(sp_or_cluster),
        'paper_claim_57_arg_proteins': 57,
        'paper_claim_self_protecting_count': 39,
        'paper_claim_more_similar_actino_count': 7,
        'paper_claim_identity_range_low': 23,
        'paper_claim_identity_range_high': 68,
    }
    write_json('C2_parser_provenance.json', result)
    return result

# ----------------------------------------------------------------------
# E-utilities helpers
# ----------------------------------------------------------------------
EUTILS = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils'

def efetch_protein_fasta(acc, retries=3, sleep=0.4):
    url = f'{EUTILS}/efetch.fcgi?db=protein&id={urllib.parse.quote(acc)}&rettype=fasta&retmode=text'
    for i in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                txt = r.read().decode()
            if txt.startswith('>'):
                return txt
        except Exception as e:
            log(f'  efetch {acc} attempt {i+1}: {e}')
            time.sleep(sleep * (i+1))
    return None

def efetch_nuccore_gb(acc, retries=3, sleep=0.5):
    url = f'{EUTILS}/efetch.fcgi?db=nuccore&id={urllib.parse.quote(acc)}&rettype=gb&retmode=text'
    for i in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                return r.read().decode()
        except Exception as e:
            log(f'  efetch nuccore {acc} attempt {i+1}: {e}')
            time.sleep(sleep * (i+1))
    return None

def esearch_protein_link_to_nuccore(acc, retries=3, sleep=0.5):
    """Return nuccore IDs linked from a protein accession."""
    # first get protein UID
    url = f'{EUTILS}/esearch.fcgi?db=protein&term={urllib.parse.quote(acc)}&retmode=json'
    try:
        with urllib.request.urlopen(url, timeout=20) as r:
            j = json.loads(r.read())
        ids = j.get('esearchresult', {}).get('idlist', [])
        if not ids:
            return []
        puid = ids[0]
    except Exception as e:
        log(f'  esearch protein {acc}: {e}')
        return []
    url2 = f'{EUTILS}/elink.fcgi?dbfrom=protein&db=nuccore&id={puid}&retmode=json'
    try:
        with urllib.request.urlopen(url2, timeout=20) as r:
            j = json.loads(r.read())
        out = []
        for ls in j.get('linksets', []):
            for ldb in ls.get('linksetdbs', []):
                if ldb.get('linkname') in ('protein_nuccore', 'protein_nuccore_wgs'):
                    out.extend(ldb.get('links', []))
        return list(dict.fromkeys(out))   # dedup, preserve order
    except Exception as e:
        log(f'  elink {acc}: {e}')
        return []

def run_blastp_pair(qfa, sfa):
    """Return (pct_identity, qcov_hsp, evalue) using local blastp."""
    import tempfile, shutil
    blastp = shutil.which('blastp')
    if not blastp:
        return None
    with tempfile.TemporaryDirectory() as td:
        q = Path(td) / 'q.fa'; q.write_text(qfa)
        s = Path(td) / 's.fa'; s.write_text(sfa)
        out = subprocess.run(
            [blastp, '-query', str(q), '-subject', str(s),
             '-outfmt', '6 pident qcovhsp evalue length nident'],
            capture_output=True, text=True, timeout=120)
        if out.returncode != 0 or not out.stdout.strip():
            return None
        line = out.stdout.strip().splitlines()[0].split('\t')
        return {
            'pident': float(line[0]),
            'qcovhsp': float(line[1]),
            'evalue': float(line[2]),
            'length': int(line[3]),
            'nident': int(line[4]),
        }

def cache_get(acc):
    """Cache protein fasta in results/repass/seqs/"""
    d = RESULTS / 'seqs'; d.mkdir(exist_ok=True)
    fp = d / f'{acc}.fa'
    if fp.exists() and fp.stat().st_size > 0:
        return fp.read_text()
    txt = efetch_protein_fasta(acc)
    if txt:
        fp.write_text(txt)
        time.sleep(0.34)   # NCBI etiquette: 3 req/sec without API key
        return txt
    return None

# ----------------------------------------------------------------------
# CLAIM C3: 7-vs-9 discrepancy in "more similar to actinobacterial" set
# ----------------------------------------------------------------------
def claim_C3_more_similar_set():
    log('C3: enumerate Y-flagged "more similar to actinobacterial" entries')
    import openpyxl
    wb = openpyxl.load_workbook(PAPER / 'supp_data1.xlsx', data_only=True)
    ws = wb['Supplementary Data 1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    y_entries = []
    for r in rows:
        if r[17] and str(r[17]).strip().upper().startswith('Y'):
            y_entries.append({
                'strep_acc': str(r[0]).strip().split()[0] if r[0] else None,
                'gene_type': r[1],
                'host': r[4],
                'proteo_acc': str(r[14]).strip().split()[0] if r[14] else None,
                'paper_ident_pct': r[13] * 100 if isinstance(r[13], (int, float)) else None,
                'isolated_from': r[18],
            })
    result = {
        'paper_text_says': 7,
        'supp_data1_y_count': len(y_entries),
        'discrepancy': len(y_entries) - 7,
        'entries': y_entries,
        'note': (
            'Paper Results §1 says "Seven of these proteobacterial proteins have sequences more '
            'similar to actinobacterial proteins than to proteins from any other phyla". '
            'Supplementary Data 1 has 9 entries marked Y in column 17. The two extra are likely '
            'near-identical paralogs of the same gene type (multiple tet/aph entries). '
            'After collapsing by proteobacterial accession the count is also re-checked below.'
        ),
    }
    uniq_proteo = sorted(set(e['proteo_acc'] for e in y_entries if e['proteo_acc']))
    result['unique_proteo_accs'] = uniq_proteo
    result['unique_proteo_count'] = len(uniq_proteo)
    write_json('C3_more_similar_set.json', result)
    return result

# ----------------------------------------------------------------------
# CLAIM C5: C. glutamicum 1014 vs Arthrobacter sp. 161MFSha2.1 cmx — 93%
# ----------------------------------------------------------------------
def claim_C5_glutamicum_vs_arthrobacter():
    log('C5: C. glutamicum 1014 cmx vs Arthrobacter sp. 161MFSha2.1 cmx — paper 93%')
    # Strategy: pull cmx from C. glutamicum 1014 via esearch; pull cmx from
    # Arthrobacter sp. 161MFSha2.1. Then BLASTP both.
    queries = [
        # Search terms — biased toward correct strain
        ('C_glutamicum_1014_cmx', 'Corynebacterium glutamicum 1014 cmx'),
        ('Arthrobacter_161MFSha21_cmx', 'Arthrobacter 161MFSha2.1 cmx'),
    ]
    out = {'paper_claim_identity_pct': 93, 'queries': []}
    seqs = {}
    for tag, term in queries:
        url = f'{EUTILS}/esearch.fcgi?db=protein&term={urllib.parse.quote(term)}&retmode=json&retmax=10'
        try:
            with urllib.request.urlopen(url, timeout=20) as r:
                j = json.loads(r.read())
            ids = j.get('esearchresult', {}).get('idlist', [])
        except Exception as e:
            ids = []
            log(f'  esearch error {term}: {e}')
        log(f'  {tag}: esearch -> {len(ids)} hits')
        # fetch first up-to-3 fasta records, pick best by header containing cmx/chloramphenicol
        fastas = []
        for uid in ids[:5]:
            url2 = f'{EUTILS}/efetch.fcgi?db=protein&id={uid}&rettype=fasta&retmode=text'
            try:
                with urllib.request.urlopen(url2, timeout=20) as r:
                    f = r.read().decode()
                fastas.append((uid, f))
                time.sleep(0.34)
            except Exception as e:
                log(f'    efetch {uid}: {e}')
        # pick one with chloramphenicol or cmx in defline
        chosen = None
        for uid, fa in fastas:
            d = fa.splitlines()[0].lower() if fa else ''
            if 'chloramphenicol' in d or ' cmx' in d or '|cmx' in d:
                chosen = (uid, fa); break
        if not chosen and fastas:
            chosen = fastas[0]
        out['queries'].append({
            'tag': tag, 'term': term, 'n_hits': len(ids),
            'chosen_uid': chosen[0] if chosen else None,
            'chosen_defline': chosen[1].splitlines()[0] if chosen else None,
        })
        if chosen:
            seqs[tag] = chosen[1]
    if len(seqs) == 2:
        b = run_blastp_pair(seqs['C_glutamicum_1014_cmx'], seqs['Arthrobacter_161MFSha21_cmx'])
        out['blastp_result'] = b
        if b:
            out['identity_match_paper'] = abs(b['pident'] - 93.0) <= 5.0
    else:
        out['blastp_result'] = None
        out['blocker'] = 'could not retrieve one or both cmx sequences from named strains'
    write_json('C5_glutamicum_vs_arthrobacter.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C6: Supp Data 4 carry-back intermediates — verify cmx in named
# carrier-sandwich genomes (C. diphtheriae BH8, C. resistens pJA144188,
# E. asburiae 35642, K. oxytoca CHS143)
# ----------------------------------------------------------------------
def claim_C6_carry_back_intermediates():
    log('C6: carry-back intermediates — confirm cmx presence in named genomes')
    targets = {
        'C_diphtheriae_BH8': 'Corynebacterium diphtheriae BH8',
        'C_resistens_pJA144188': 'pJA144188 Corynebacterium resistens',
        'E_asburiae_35642': 'Enterobacter asburiae 35642',
        'K_oxytoca_CHS143': 'Klebsiella oxytoca CHS143',
    }
    res = {}
    for tag, term in targets.items():
        # search nuccore for the genome/plasmid; look for cmx in features
        url = f'{EUTILS}/esearch.fcgi?db=nuccore&term={urllib.parse.quote(term)}&retmode=json&retmax=5'
        try:
            with urllib.request.urlopen(url, timeout=20) as r:
                j = json.loads(r.read())
            ids = j.get('esearchresult', {}).get('idlist', [])
        except Exception as e:
            ids = []
            log(f'  esearch {term}: {e}')
        log(f'  {tag}: nuccore -> {len(ids)} hits')
        # search for cmx-containing proteins directly via protein db too
        url2 = f'{EUTILS}/esearch.fcgi?db=protein&term={urllib.parse.quote(term + " AND chloramphenicol")}&retmode=json&retmax=5'
        try:
            with urllib.request.urlopen(url2, timeout=20) as r:
                j2 = json.loads(r.read())
            pids = j2.get('esearchresult', {}).get('idlist', [])
        except Exception:
            pids = []
        deflines = []
        for pid in pids[:3]:
            try:
                with urllib.request.urlopen(f'{EUTILS}/efetch.fcgi?db=protein&id={pid}&rettype=fasta&retmode=text', timeout=15) as r:
                    f = r.read().decode()
                deflines.append(f.splitlines()[0])
                time.sleep(0.34)
            except Exception:
                pass
        res[tag] = {
            'search_term': term,
            'nuccore_hits': len(ids), 'nuccore_first_ids': ids[:3],
            'protein_chloramphenicol_hits': len(pids), 'protein_deflines': deflines,
            'cmx_evidence': any('chloramphenicol' in d.lower() or 'cmx' in d.lower() for d in deflines),
        }
    out = {
        'paper_claim': (
            'Carry-back intermediates with cmx + carrier-sandwich found in '
            'C. diphtheriae BH8, C. resistens pJA144188, E. asburiae 35642, K. oxytoca CHS143'
        ),
        'per_target': res,
        'note': (
            'Full sandwich-structure verification requires parsing GenBank features '
            'around cmx and computing distances to IS6100/orf5/sul1; here we record '
            'presence/absence of cmx-family chloramphenicol-resistance protein in each '
            'named genome and report the genome accession candidates for future synteny '
            'reproduction.'
        ),
    }
    write_json('C6_carry_back_intermediates.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C8: Cmx >99% identical to non-Streptomyces actinobacteria
# (already verified by pass1 via BV-BRC; ground-test via NCBI BLAST top hits)
# ----------------------------------------------------------------------
def claim_C8_cmx_99pct_actino():
    log('C8: Cmx (WP_005297378.1) >99% to non-Streptomyces actinobacteria')
    cmx = cache_get('WP_005297378.1')
    # Pull top homologs from a few non-Streptomyces actino genera: Corynebacterium, Arthrobacter, Microbacterium
    test_accs = [
        # representative Corynebacterium cmx
        'WP_010935478.1',     # C. diphtheriae chloramphenicol exporter (well-known)
        'WP_011544419.1',     # C. resistens cmx
        'WP_005536488.1',     # C. glutamicum cmx
    ]
    pairs = []
    for a in test_accs:
        fa = cache_get(a)
        if not fa or not cmx:
            pairs.append({'subject': a, 'error': 'fetch_failed'})
            continue
        b = run_blastp_pair(cmx, fa)
        pairs.append({
            'subject': a,
            'subject_defline': fa.splitlines()[0],
            'blastp': b,
            'meets_99pct': bool(b and b['pident'] >= 99.0),
        })
    matches_paper = any(p.get('meets_99pct') for p in pairs)
    out = {
        'query_acc': 'WP_005297378.1 (proteobacterial Cmx)',
        'paper_claim': 'gene identical or nearly identical (>99% identity) to non-Streptomyces actinobacteria',
        'tested_subjects': pairs,
        'any_subject_at_99pct_or_higher': matches_paper,
    }
    write_json('C8_cmx_99pct_actino.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C9: cmx + tnp45 transposon colocalization in named genomes
# ----------------------------------------------------------------------
def claim_C9_cmx_tnp45_synteny():
    log('C9: cmx + tnp45 transposon colocalization')
    # Approach: search NCBI nuccore for "cmx tnp45" or "chloramphenicol transposase tnp45"
    # then count how many hit records appear; download a GenBank for one and confirm
    # both gene names appear within ~5kb of each other.
    url = f'{EUTILS}/esearch.fcgi?db=nuccore&term=cmx+AND+tnp45&retmode=json&retmax=20'
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            j = json.loads(r.read())
        ids = j.get('esearchresult', {}).get('idlist', [])
        total = int(j.get('esearchresult', {}).get('count', 0))
    except Exception as e:
        ids, total = [], 0
        log(f'  esearch cmx+tnp45: {e}')
    log(f'  cmx+tnp45 nuccore search: {total} total records, sampled {len(ids)}')
    sample_genbank = None; sample_id = None; colocated = False; dist = None
    if ids:
        sample_id = ids[0]
        gb = efetch_nuccore_gb(sample_id)
        if gb:
            (RESULTS / 'sample_cmx_tnp45.gb').write_text(gb)
            # quick proximity check
            import re
            # find feature ranges for cmx and tnp45
            cmx_pos = []
            tnp_pos = []
            for m in re.finditer(r'/gene="([^"]+)"', gb):
                pass
            # simpler: look for "cmx" and "tnp45" lines and grab the previous CDS line
            lines = gb.splitlines()
            for i, l in enumerate(lines):
                low = l.lower()
                if '/gene="cmx' in low or '/product="chloramphenicol' in low:
                    # backtrack to nearest CDS/feature line
                    for j in range(i, max(0, i-15), -1):
                        if lines[j].lstrip().startswith(('CDS ', 'gene ', 'mobile_element')):
                            mm = re.search(r'(\d+)\.\.(\d+)', lines[j])
                            if mm: cmx_pos.append((int(mm.group(1)), int(mm.group(2)))); break
                if '/gene="tnp45' in low or 'tnp45' in low.replace('"','').replace("'",''):
                    for j in range(i, max(0, i-15), -1):
                        if lines[j].lstrip().startswith(('CDS ', 'gene ', 'mobile_element')):
                            mm = re.search(r'(\d+)\.\.(\d+)', lines[j])
                            if mm: tnp_pos.append((int(mm.group(1)), int(mm.group(2)))); break
            if cmx_pos and tnp_pos:
                # min midpoint distance
                def mid(p): return (p[0]+p[1])//2
                dist = min(abs(mid(c)-mid(t)) for c in cmx_pos for t in tnp_pos)
                colocated = dist <= 5000
            sample_genbank = lines[0][:80] if lines else None
    out = {
        'paper_claim': 'cmx is colocalized with actinobacterial transposase gene tnp45 forming a transposon',
        'ncbi_nuccore_cmx_AND_tnp45_total_hits': total,
        'sampled_id_count': len(ids),
        'sampled_first_id': sample_id,
        'sample_locus_line': sample_genbank,
        'cmx_tnp45_colocation_distance_bp': dist,
        'colocated_within_5kb': colocated,
    }
    write_json('C9_cmx_tnp45_synteny.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C10: 9 environmental, 3 pathogen split of the 12 HGT proteins
# ----------------------------------------------------------------------
def claim_C10_env_vs_pathogen_split():
    log('C10: 9 environmental vs 3 pathogen split among 12 HGT proteobacterial proteins')
    # Use Supp Data 1 "isolated from" column (18). Filter to Y-marked + the 3 known
    # pathogens called out in text (APH(3"), and the 2 recent transfers cmx + lmrA).
    import openpyxl
    wb = openpyxl.load_workbook(PAPER / 'supp_data1.xlsx', data_only=True)
    ws = wb['Supplementary Data 1']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    # Build 12-HGT set: union of Y-flagged + Cmx P31141 + LmrA CAA42550 + APH WP_031942890.1 line + Sul1 etc.
    # Pragmatic approach: use Y-flagged set + supplement with the explicit ones the paper names
    entries = []
    for r in rows:
        is_y = r[17] and str(r[17]).strip().upper().startswith('Y')
        proteo = str(r[14]).strip().split()[0] if r[14] else None
        if is_y or proteo in ('WP_031942890.1','WP_005297378.1','WP_038989331.1'):
            entries.append({
                'strep_acc': str(r[0]).strip().split()[0] if r[0] else None,
                'gene_type': r[1],
                'proteo_acc': proteo,
                'isolated_from': r[18],
            })
    # de-dup on proteo_acc
    seen = set(); uniq = []
    for e in entries:
        if e['proteo_acc'] in seen: continue
        seen.add(e['proteo_acc']); uniq.append(e)

    pathogen_keywords = ['patient', 'clinical', 'farm animal', 'human', 'infection', 'hospital']
    env_keywords = ['soil', 'environment', 'water', 'sediment', 'plant', 'compost']
    for e in uniq:
        s = (e['isolated_from'] or '').lower()
        if any(k in s for k in pathogen_keywords):
            e['category'] = 'pathogen'
        elif any(k in s for k in env_keywords):
            e['category'] = 'environmental'
        else:
            e['category'] = 'unclassified'
    n_path = sum(1 for e in uniq if e['category']=='pathogen')
    n_env = sum(1 for e in uniq if e['category']=='environmental')
    n_unc = sum(1 for e in uniq if e['category']=='unclassified')
    out = {
        'paper_claim': '9 of the 12 HGT proteobacterial proteins encoded in environmental species, 3 in pathogens',
        'detected_n_entries': len(uniq),
        'pathogen': n_path,
        'environmental': n_env,
        'unclassified': n_unc,
        'matches_paper_9_env_3_pathogen': (n_env == 9 and n_path == 3),
        'entries': uniq,
        'note': (
            'Supplementary Data 1 has only sparse "isolated from" annotation, so most '
            'entries fall to unclassified. The paper categorization may rely on PATRIC/external '
            'metadata; we record this as a coverage limitation rather than a contradiction.'
        ),
    }
    write_json('C10_env_vs_pathogen_split.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C11: sul1 AFN41071.1 vs ALJ92876.1 — re-blast under repass provenance
# ----------------------------------------------------------------------
def claim_C11_sul1_reverification():
    log('C11: sul1 cross-phylum 95% identity re-verification')
    q = cache_get('AFN41071.1')
    s = cache_get('ALJ92876.1')
    out = {'paper_claim_pct': 95}
    if q and s:
        b = run_blastp_pair(q, s)
        out['blastp_result'] = b
        if b:
            out['delta_from_paper'] = b['pident'] - 95.0
            out['matches_paper'] = abs(b['pident'] - 95.0) <= 2.0
    else:
        out['error'] = 'fetch_failed'
    write_json('C11_sul1_reverification.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C12: LmrA carrier — protein WP_038989331.1 located on RSF1010-like plasmid
# ----------------------------------------------------------------------
def claim_C12_lmra_rsf1010():
    log('C12: LmrA WP_038989331.1 carrier plasmid check (paper: RSF1010-like)')
    nucs = esearch_protein_link_to_nuccore('WP_038989331.1')
    log(f'  LmrA linked nuccore records: {len(nucs)}')
    sample_meta = []
    plasmid_hits = 0
    rsf_hits = 0
    for uid in nucs[:15]:
        url = f'{EUTILS}/esummary.fcgi?db=nuccore&id={uid}&retmode=json'
        try:
            with urllib.request.urlopen(url, timeout=15) as r:
                j = json.loads(r.read())
            doc = j['result'][uid]
            title = (doc.get('title') or '')
            sample_meta.append({'uid': uid, 'title': title[:160]})
            tl = title.lower()
            if 'plasmid' in tl: plasmid_hits += 1
            if 'rsf1010' in tl or 'rsf 1010' in tl: rsf_hits += 1
            time.sleep(0.34)
        except Exception as e:
            log(f'    esummary {uid}: {e}')
    out = {
        'paper_claim': 'WP_038989331.1 (LmrA) located on an RSF1010-like plasmid',
        'linked_nuccore_total': len(nucs),
        'sampled_meta': sample_meta,
        'sampled_n': len(sample_meta),
        'plasmid_titles_in_sample': plasmid_hits,
        'rsf1010_titles_in_sample': rsf_hits,
        'plasmid_evidence': plasmid_hits > 0,
        'rsf1010_evidence': rsf_hits > 0,
    }
    write_json('C12_lmra_rsf1010.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C13: APH(3'') WP_031942890.1 in pathogens — verify host annotations
# ----------------------------------------------------------------------
def claim_C13_aph3_in_pathogens():
    log('C13: APH(3") WP_031942890.1 host pathogen verification')
    fa = cache_get('WP_031942890.1')
    nucs = esearch_protein_link_to_nuccore('WP_031942890.1')
    sample = []
    pathogen_titles = 0
    for uid in nucs[:20]:
        url = f'{EUTILS}/esummary.fcgi?db=nuccore&id={uid}&retmode=json'
        try:
            with urllib.request.urlopen(url, timeout=15) as r:
                j = json.loads(r.read())
            doc = j['result'][uid]
            title = (doc.get('title') or '')[:160]
            organism = doc.get('organism') or ''
            sample.append({'uid': uid, 'title': title, 'organism': organism})
            tl = (organism + ' ' + title).lower()
            if any(p in tl for p in ['salmonella','pseudomonas','klebsiella','e. coli','escherichia','enterobacter','acinetobacter','vibrio']):
                pathogen_titles += 1
            time.sleep(0.34)
        except Exception as e:
            log(f'    esummary {uid}: {e}')
    out = {
        'paper_claim': 'APH(3″) WP_031942890.1 is harboured by pathogens',
        'linked_nuccore_total': len(nucs),
        'sampled_meta': sample,
        'pathogen_genus_hits': pathogen_titles,
        'pathogen_evidence': pathogen_titles > 0,
        'defline': fa.splitlines()[0] if fa else None,
    }
    write_json('C13_aph3_in_pathogens.json', out)
    return out

# ----------------------------------------------------------------------
# CLAIM C14: cmx-tnp45 transposon distribution in actinobacteria + proteobacteria
# (number of distinct genera with both)
# ----------------------------------------------------------------------
def claim_C14_cmx_tnp45_distribution():
    log('C14: cmx+tnp45 distribution across genera (actino vs proteo)')
    # Search for cmx+tnp45 in nuccore, fetch top 30 summaries, group by organism
    url = f'{EUTILS}/esearch.fcgi?db=nuccore&term=cmx+tnp45&retmode=json&retmax=50'
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            j = json.loads(r.read())
        ids = j.get('esearchresult', {}).get('idlist', [])
    except Exception as e:
        ids = []
    genera_actino = set(); genera_proteo = set(); other = set()
    ACTINO = {'streptomyces','corynebacterium','arthrobacter','mycobacterium','microbacterium','rhodococcus','nocardia','actinomyces','paenarthrobacter','curtobacterium'}
    PROTEO = {'pseudomonas','escherichia','klebsiella','enterobacter','acinetobacter','salmonella','vibrio','burkholderia','aeromonas','citrobacter','proteus','serratia'}
    docs = []
    for uid in ids[:50]:
        url2 = f'{EUTILS}/esummary.fcgi?db=nuccore&id={uid}&retmode=json'
        try:
            with urllib.request.urlopen(url2, timeout=15) as r:
                jj = json.loads(r.read())
            doc = jj['result'][uid]
            org = (doc.get('organism') or '').lower()
            gen = org.split()[0] if org else ''
            docs.append({'uid': uid, 'organism': doc.get('organism'), 'title': (doc.get('title') or '')[:120]})
            if gen in ACTINO: genera_actino.add(gen)
            elif gen in PROTEO: genera_proteo.add(gen)
            elif gen: other.add(gen)
            time.sleep(0.34)
        except Exception:
            pass
    out = {
        'paper_claim': 'intact cmx+tnp45 transposon found in both actinobacteria and proteobacteria',
        'sampled_n': len(docs),
        'actino_genera_with_cmx_tnp45': sorted(genera_actino),
        'proteo_genera_with_cmx_tnp45': sorted(genera_proteo),
        'other_genera_with_cmx_tnp45': sorted(other),
        'both_phyla_present': bool(genera_actino) and bool(genera_proteo),
        'docs_sample': docs[:20],
    }
    write_json('C14_cmx_tnp45_distribution.json', out)
    return out


def main():
    summary = {'run_started': time.strftime('%Y-%m-%d %H:%M:%S')}
    log('== RE-PASS START ==')
    for name, fn in [
        ('C2_parser_provenance', claim_C2_parser_provenance),
        ('C3_more_similar_set', claim_C3_more_similar_set),
        ('C11_sul1_reverification', claim_C11_sul1_reverification),
        ('C8_cmx_99pct_actino', claim_C8_cmx_99pct_actino),
        ('C13_aph3_in_pathogens', claim_C13_aph3_in_pathogens),
        ('C12_lmra_rsf1010', claim_C12_lmra_rsf1010),
        ('C9_cmx_tnp45_synteny', claim_C9_cmx_tnp45_synteny),
        ('C14_cmx_tnp45_distribution', claim_C14_cmx_tnp45_distribution),
        ('C5_glutamicum_vs_arthrobacter', claim_C5_glutamicum_vs_arthrobacter),
        ('C6_carry_back_intermediates', claim_C6_carry_back_intermediates),
        ('C10_env_vs_pathogen_split', claim_C10_env_vs_pathogen_split),
    ]:
        try:
            summary[name] = fn()
        except Exception as e:
            log(f'!! {name} failed: {e}')
            import traceback; traceback.print_exc()
            summary[name] = {'error': str(e)}
    summary['run_finished'] = time.strftime('%Y-%m-%d %H:%M:%S')
    write_json('repass_summary.json', summary)
    log('== RE-PASS DONE ==')

if __name__ == '__main__':
    main()
