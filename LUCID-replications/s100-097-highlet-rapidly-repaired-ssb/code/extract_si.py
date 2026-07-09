"""Extract Supplementary Excel into clean CSVs for downstream analysis.

Source: 41598_2023_38295_MOESM2_ESM.xlsx
Produces:
    data/foci.csv         (Figure 1 53BP1 foci kinetics)
    data/clonogenic.csv   (Figure 2 dose-response)
    data/sld.csv          (Figure 3 SLD repair / interval-survival)
"""
from pathlib import Path
import csv
import openpyxl

HERE = Path(__file__).resolve().parent
PROJ = HERE.parent
SRC = PROJ / "source" / "SI_MOESM2_ESM.xlsx"
DATA = PROJ / "code" / "data"
DATA.mkdir(parents=True, exist_ok=True)


def num(x):
    if x is None or x == "" or x == "-":
        return None
    try:
        return float(x)
    except Exception:
        return None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # ---- Foci (Figure 1) ----
    ws = wb["Supp Data 1 - Figure 1"]
    rows = list(ws.iter_rows(values_only=True))
    foci = []
    cell_line = None
    for r in rows:
        if r and r[1] and isinstance(r[1], str) and r[1].strip() in ("U2OS", "PC-3"):
            cell_line = r[1].strip()
            continue
        if cell_line and r and r[1] is not None:
            t = num(r[1])
            if t is None:
                continue
            xr_mean, xr_sd = num(r[2]), num(r[3])
            a_mean, a_sd = num(r[4]), num(r[5])
            foci.append((cell_line, t, xr_mean, xr_sd, a_mean, a_sd))
    with (DATA / "foci.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["cell_line", "time_h", "xray_mean", "xray_sd",
                    "alpha_mean", "alpha_sd"])
        w.writerows(foci)

    # ---- Clonogenic (Figure 2) ----
    ws = wb["Supp Data 2 - Figure 2"]
    rows = list(ws.iter_rows(values_only=True))
    clono = []
    cell_line = None
    for r in rows:
        if r and r[1] and isinstance(r[1], str) and r[1].strip() in ("U2OS", "PC-3"):
            cell_line = r[1].strip()
            continue
        if cell_line and r and r[1] is not None:
            d = num(r[1])
            if d is None:
                continue
            add_m = num(r[2])
            xr_m, xr_sd = num(r[3]), num(r[4])
            a_m, a_sd = num(r[5]), num(r[6])
            xa_m, xa_sd = num(r[7]), num(r[8])
            ax_m, ax_sd = num(r[9]), num(r[10])
            clono.append((cell_line, d, add_m, xr_m, xr_sd, a_m, a_sd,
                          xa_m, xa_sd, ax_m, ax_sd))
    with (DATA / "clonogenic.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["cell_line", "total_dose_Gy", "additive_model",
                    "xray_mean", "xray_sd",
                    "alpha_mean", "alpha_sd",
                    "xa_mean", "xa_sd",
                    "ax_mean", "ax_sd"])
        w.writerows(clono)

    # ---- SLD repair (Figure 3) ----
    ws = wb["Supp Data 3 - Figure 3"]
    rows = list(ws.iter_rows(values_only=True))
    sld = []
    cell_line = None
    for r in rows:
        if r and r[1] and isinstance(r[1], str) and r[1].strip() in ("U2OS", "PC-3"):
            cell_line = r[1].strip()
            continue
        if cell_line and r and r[1] is not None:
            t = num(r[1])
            if t is None:
                continue
            xx_m, xx_sd = num(r[2]), num(r[3])
            aa_m, aa_sd = num(r[4]), num(r[5])
            xa_m, xa_sd = num(r[6]), num(r[7])
            ax_m, ax_sd = num(r[8]), num(r[9])
            sld.append((cell_line, t, xx_m, xx_sd, aa_m, aa_sd,
                        xa_m, xa_sd, ax_m, ax_sd))
    with (DATA / "sld.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["cell_line", "interval_h",
                    "xx_mean", "xx_sd",
                    "aa_mean", "aa_sd",
                    "xa_mean", "xa_sd",
                    "ax_mean", "ax_sd"])
        w.writerows(sld)

    print(f"wrote {DATA}/foci.csv ({len(foci)} rows)")
    print(f"wrote {DATA}/clonogenic.csv ({len(clono)} rows)")
    print(f"wrote {DATA}/sld.csv ({len(sld)} rows)")


if __name__ == "__main__":
    main()
