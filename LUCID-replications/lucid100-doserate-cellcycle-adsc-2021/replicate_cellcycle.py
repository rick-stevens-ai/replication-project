"""
Independent re-derivation of cell cycle statistics from Rusin et al. 2021
(PLoS ONE 16(4):e0250160), Mendeley data DOI 10.17632/8t594k4w8z.

We re-compute:
  * per-condition (Control/LDR/HDR) means and SDs of G0/G1, S, G2/M
    phase percentages at 12h, Day 1, Day 2, Day 3
  * one-way ANOVA per (timepoint, phase) across {Control, LDR, HDR}
  * pairwise t-tests (Welch) with p<0.05 flag, matching the paper's
    description of which contrasts are reported significant
  * compare our values to the paper's table-style summary inside Figure3xlsx
    (Sheet1, "Mean / Std. Dev." block)

Run: python3 replicate_cellcycle.py
"""

import json
import openpyxl
import numpy as np
from scipy import stats

XLSX = "mendeley_data/Figure3Cell_Cycle_Analysis.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Sheet1"]
rows = [list(r) for r in ws.iter_rows(values_only=True)]

# Per-replicate block: rows index 3..14 (0-based), 12 columns of phase data after label
# Column layout (after label col 0):
#  cols 1..3   -> 12h:  G0/G1, S, G2/M
#  cols 4..6   -> Day 1
#  cols 7..9   -> Day 2
#  cols 10..12 -> Day 3
# NB: rows are 0-indexed from iter_rows; per inspection labels live at rows 2..13.
labels_groups = {
    "Control": [2, 3, 4, 5],   # Control 1..4 (12h only has 4; later timepoints only Control 1..3 valid)
    "LDR":     [6, 7, 8, 9],
    "HDR":     [10, 11, 12, 13],
}
timepoints = {
    "12h":  (1, 2, 3),
    "Day1": (4, 5, 6),
    "Day2": (7, 8, 9),
    "Day3": (10, 11, 12),
}
phases = ["G0/G1", "S", "G2/M"]

def get_value(rows, ridx, cidx):
    try:
        v = rows[ridx][cidx]
    except IndexError:
        return None
    return v

def collect(condition, tp):
    cols = timepoints[tp]
    out = [[], [], []]  # one list per phase
    for ridx in labels_groups[condition]:
        r = rows[ridx]
        if not r or not isinstance(r[0], str):
            continue
        # only first 3 replicates are real for Day1..Day3 (per data);
        # 12h has 4 replicates for each condition.
        vals = [r[cols[0]], r[cols[1]], r[cols[2]]]
        # skip the all-zero row used as "missing" marker (LDR 3 at Day 3 -> 0/0/0)
        if all((v == 0 or v is None) for v in vals):
            continue
        # skip rows that are blank for this timepoint (Day1..Day3 for replicate 4)
        if all(v is None for v in vals):
            continue
        for k, v in enumerate(vals):
            if v is None:
                continue
            out[k].append(float(v))
    return out  # list of three lists [G0/G1 vals, S vals, G2/M vals]

# Paper-reported means/SDs are in rows 18..23 of the same sheet
# Row layout:
# row 18: header  Day1 Day2 Day3 across columns 3..11
# row 19: phase labels (G0/G1, S, G2/M repeated)
# row 20: 'Control Mean'
# row 21: 'Std. Dev.'
# row 22: 'LDR Mean'
# row 23: 'Std. Dev.'  -> note: there's also 'HDR Mean' lower; the dump showed it
# Re-read raw layout to be safe
print("=== Re-derived stats from per-replicate data ===")
recomputed = {}
for tp in timepoints:
    recomputed[tp] = {}
    for cond in labels_groups:
        recomputed[tp][cond] = {}
        cols = collect(cond, tp)
        for ph_idx, ph in enumerate(phases):
            vals = cols[ph_idx]
            if vals:
                recomputed[tp][cond][ph] = {
                    "n": len(vals),
                    "mean": float(np.mean(vals)),
                    "sd": float(np.std(vals, ddof=1)) if len(vals) > 1 else float("nan"),
                    "values": vals,
                }
            else:
                recomputed[tp][cond][ph] = {"n": 0, "mean": None, "sd": None, "values": []}

# Print table
for tp in timepoints:
    print(f"\n-- {tp} --")
    print(f"  {'Cond':<8} {'Phase':<6} {'n':>2}  {'mean':>7}  {'sd':>7}  values")
    for cond in labels_groups:
        for ph in phases:
            v = recomputed[tp][cond][ph]
            if v["n"]:
                vs = ",".join(f"{x:.2f}" for x in v["values"])
                print(f"  {cond:<8} {ph:<6} {v['n']:>2}  {v['mean']:7.3f}  {v['sd'] if v['sd'] is not None else float('nan'):7.3f}  [{vs}]")
            else:
                print(f"  {cond:<8} {ph:<6} {0:>2}    --       --   []")

# Paper's reported means/SDs (Day1/Day2/Day3 only) from the worksheet's summary block,
# verbatim from sheet dump:
paper_summary = {
    # tp -> cond -> phase -> (mean, sd)
    "Day1": {
        "Control": {"G0/G1": (65.97, 1.559),  "S": (15.42, 3.132),  "G2/M": (17.97, 3.179)},
        "LDR":     {"G0/G1": (67.40, 1.240),  "S": (7.963, 2.221),  "G2/M": (24.44, 2.372)},
        "HDR":     {"G0/G1": (61.20, 1.838),  "S": (9.853, 0.4163), "G2/M": (28.70, 1.397)},
    },
    "Day2": {
        "Control": {"G0/G1": (67.28, 3.850),  "S": (5.943, 0.4827), "G2/M": (26.74, 3.467)},
        "LDR":     {"G0/G1": (63.32, 1.018),  "S": (3.783, 0.5320), "G2/M": (32.82, 1.477)},
        "HDR":     {"G0/G1": (59.28, 1.425),  "S": (5.837, 0.3755), "G2/M": (34.87, 1.107)},
    },
    "Day3": {
        "Control": {"G0/G1": (68.06, 3.632),  "S": (7.980, 1.520),  "G2/M": (23.94, 2.118)},
        # LDR Day3 in raw rows: rep1=62.13,8.03,29.81; rep2=60.79,9.10,30.08; rep3=0,0,0 (missing flag)
        "LDR":     {"G0/G1": (61.46, 0.9475), "S": (8.565, 0.7566), "G2/M": (29.95, 0.1909)},
        "HDR":     {"G0/G1": (57.06, 0.8402), "S": (8.000, 0.378),  "G2/M": (34.87, 1.132)},
    },
}

print("\n=== Paper-reported vs Re-derived (Day1/Day2/Day3) ===")
print(f"  {'TP':<5} {'Cond':<8} {'Phase':<6}  {'paper_mean':>10} {'paper_sd':>9}  {'repl_mean':>10} {'repl_sd':>9}  Δmean    ok?")
tolerance_mean = 0.05  # absolute pct units
tolerance_sd   = 0.05
mismatches = []
matches = 0
total = 0
for tp in ("Day1", "Day2", "Day3"):
    for cond in ("Control", "LDR", "HDR"):
        for ph in phases:
            pm, ps = paper_summary[tp][cond][ph]
            rv = recomputed[tp][cond][ph]
            rm, rs = rv["mean"], rv["sd"]
            total += 1
            d_mean = (rm - pm) if rm is not None else None
            d_sd   = (rs - ps) if rs is not None else None
            ok = (rm is not None and rs is not None
                  and abs(d_mean) <= tolerance_mean
                  and abs(d_sd) <= tolerance_sd)
            if ok:
                matches += 1
            else:
                mismatches.append((tp, cond, ph, pm, ps, rm, rs, d_mean, d_sd))
            print(f"  {tp:<5} {cond:<8} {ph:<6}  {pm:>10.3f} {ps:>9.3f}  {rm if rm is not None else float('nan'):>10.3f} {rs if rs is not None else float('nan'):>9.3f}  {d_mean if d_mean is not None else float('nan'):>+6.3f}   {'OK' if ok else 'DIFF'}")

print(f"\nSummary table-match: {matches}/{total} cells matched within ±{tolerance_mean} (mean) / ±{tolerance_sd} (sd)")
if mismatches:
    print("Mismatches:")
    for m in mismatches:
        print("  ", m)

# Statistical tests: per (timepoint, phase) across Control/LDR/HDR
#   one-way ANOVA + pairwise Welch t-tests
print("\n=== Statistical tests (re-derived) ===")
sig_alpha = 0.05
contrasts = [("Control","LDR"), ("Control","HDR"), ("LDR","HDR")]

# Paper-reported significance pattern (from Results section + Fig 3 caption):
# Fig 3b (G0/G1): HDR significantly different from both Control and LDR at ALL time points;
#                 LDR significantly different from Control at Day 3 only.
# Fig 3c (S):     LDR and HDR significantly different from Control at Day 1;
#                 LDR different from both Control AND HDR at Day 2;
#                 no differences at Day 3.
# Fig 3d (G2/M):  LDR and HDR each significantly different from Control at all time points;
#                 LDR and HDR significantly different from each other at Day 3.

paper_expected = {
    # (tp, phase): { contrast: True/False }
    ("Day1","G0/G1"): {("Control","LDR"): False, ("Control","HDR"): True,  ("LDR","HDR"): True},
    ("Day2","G0/G1"): {("Control","LDR"): False, ("Control","HDR"): True,  ("LDR","HDR"): True},
    ("Day3","G0/G1"): {("Control","LDR"): True,  ("Control","HDR"): True,  ("LDR","HDR"): True},
    ("Day1","S"):     {("Control","LDR"): True,  ("Control","HDR"): True,  ("LDR","HDR"): False},
    ("Day2","S"):     {("Control","LDR"): True,  ("Control","HDR"): False, ("LDR","HDR"): True},
    ("Day3","S"):     {("Control","LDR"): False, ("Control","HDR"): False, ("LDR","HDR"): False},
    ("Day1","G2/M"):  {("Control","LDR"): True,  ("Control","HDR"): True,  ("LDR","HDR"): False},
    ("Day2","G2/M"):  {("Control","LDR"): True,  ("Control","HDR"): True,  ("LDR","HDR"): False},
    ("Day3","G2/M"):  {("Control","LDR"): True,  ("Control","HDR"): True,  ("LDR","HDR"): True},
}

print(f"\n  {'TP':<5} {'Phase':<6} {'ANOVA p':>10}  contrast               ttest_p   sig?  paper-sig?  agree?")
agree = 0
checked = 0
for tp in ("Day1","Day2","Day3"):
    for ph in phases:
        groups = []
        for cond in ("Control","LDR","HDR"):
            groups.append(recomputed[tp][cond][ph]["values"])
        if any(len(g) < 2 for g in groups):
            anova_p = float("nan")
        else:
            anova_p = stats.f_oneway(*groups).pvalue
        for c in contrasts:
            a = recomputed[tp][c[0]][ph]["values"]
            b = recomputed[tp][c[1]][ph]["values"]
            if len(a) < 2 or len(b) < 2:
                tp_p = float("nan")
                sig = None
            else:
                tp_p = stats.ttest_ind(a, b, equal_var=False).pvalue
                sig = tp_p < sig_alpha
            exp = paper_expected[(tp,ph)][c]
            ag = (sig == exp)
            checked += 1
            if ag:
                agree += 1
            print(f"  {tp:<5} {ph:<6} {anova_p:>10.4f}  {c[0]:>7}-vs-{c[1]:<7}  {tp_p:>8.4f}  {'YES' if sig else 'no ':<4}  {'YES' if exp else 'no ':<10}  {'OK' if ag else 'DIFF'}")

print(f"\nStatistical-pattern agreement: {agree}/{checked} contrasts matched paper-reported significance")

# Save a JSON of the recomputed table for the report
out = {
    "tolerance_mean_pct": tolerance_mean,
    "tolerance_sd_pct":   tolerance_sd,
    "recomputed": recomputed,
    "paper_summary": paper_summary,
    "table_cells_matched": matches,
    "table_cells_total":   total,
    "stat_contrasts_matched": agree,
    "stat_contrasts_total":   checked,
}
with open("replication_cellcycle_results.json", "w") as fh:
    json.dump(out, fh, indent=2, default=lambda x: None if x is None else float(x))
print("\nWrote replication_cellcycle_results.json")
