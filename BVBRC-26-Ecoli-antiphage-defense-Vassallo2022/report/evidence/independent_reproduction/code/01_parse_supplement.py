#!/usr/bin/env python3
"""Independent parse of Vassallo 2022 supplementary xlsx.

Re-derives (from scratch, without touching the replication's parsed JSONs):
  - Table S5 (source strains): count, GCA accessions
  - Table S2 (systems): 21 defence systems, 32 protein components,
    source strain, contig accession, coords, protein GenBank accessions
  - Table S4 (Gao 2020 comparison): novelty numbers
"""
import json, sys, re
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parents[3].parent / "work" / "SupplementaryTables.xlsx"
OUT = Path(__file__).resolve().parents[1] / "data"
OUT.mkdir(parents=True, exist_ok=True)

wb = load_workbook(XLSX, data_only=True, read_only=True)
print("Sheet names:", wb.sheetnames)

def dump_rows(sheet, max_rows=5):
    rows = list(sheet.iter_rows(values_only=True))
    print(f"[{sheet.title}] rows={len(rows)}")
    for r in rows[:max_rows]:
        print("  ", r)
    return rows

# ---- Table S5: source strains ----
# Find sheet whose header/content mentions ECOR + assembly accession
s5_rows = None
s5_name = None
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: continue
    text = " ".join(str(c) for r in rows[:3] for c in r if c is not None).lower()
    if "assembly" in text and ("ecor" in text or "genbank" in text or "acquired" in text):
        s5_rows = rows
        s5_name = name
        break

print(f"\n--- S5 candidate sheet: {s5_name} ---")
if s5_rows:
    for r in s5_rows[:3]:
        print("  hdr:", r)

# Try to identify header row + extract assembly accession column
def find_col(header, patterns):
    for i, h in enumerate(header):
        if h is None: continue
        hl = str(h).strip().lower()
        for pat in patterns:
            if pat in hl:
                return i
    return None

if s5_rows:
    # try header at row 0 or 1
    for hdr_idx in [0, 1]:
        header = s5_rows[hdr_idx]
        asm_col = find_col(header, ["assembly", "gca"])
        name_col = find_col(header, ["strain", "name", "genome"])
        if asm_col is not None:
            break
    print(f"header row idx {hdr_idx}, asm_col={asm_col}, name_col={name_col}")
    body = s5_rows[hdr_idx+1:]
    strains = []
    for r in body:
        if not r or all(c is None for c in r): continue
        asm = r[asm_col] if asm_col is not None and asm_col < len(r) else None
        nm = r[name_col] if name_col is not None and name_col < len(r) else None
        if not asm: continue
        asm_s = str(asm).strip()
        m = re.match(r"(GC[AF]_\d+\.\d+)", asm_s)
        if m:
            strains.append({"name": nm, "assembly": m.group(1)})
    print(f"S5 strain count: {len(strains)}")
    with open(OUT / "indep_s5_strains.json", "w") as fh:
        json.dump(strains, fh, indent=2)

# ---- Table S2: systems ----
# heuristic: find a sheet whose content includes "PD-T4" and contig accessions like QOX*/RRW*
s2_rows = None; s2_name = None
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    text_all = " ".join(str(c) for r in rows[:60] for c in r if c is not None)
    if "PD-T4-1" in text_all or "PD-\u03bb" in text_all or "PD-λ" in text_all or "PD-T7" in text_all:
        s2_rows = rows; s2_name = name; break
print(f"\n--- S2 candidate sheet: {s2_name} ---")
if s2_rows:
    for r in s2_rows[:6]:
        print("  ", r)

if s2_rows:
    # find header row containing "system" or "source strain" or "protein"
    hdr_idx = None
    for i, r in enumerate(s2_rows[:10]):
        if not r: continue
        txt = " ".join(str(c) for c in r if c is not None).lower()
        if ("system" in txt or "pd" in txt) and ("source" in txt or "strain" in txt or "protein" in txt or "contig" in txt):
            hdr_idx = i; break
    if hdr_idx is None:
        hdr_idx = 0
    header = s2_rows[hdr_idx]
    print(f"S2 header row {hdr_idx}: {header}")
    sys_col = find_col(header, ["system", "pd"])
    src_col = find_col(header, ["source strain", "strain", "source"])
    contig_col = find_col(header, ["contig", "accession"])
    prot_col = find_col(header, ["protein", "cds", "genbank", "accession"])
    coord_col = find_col(header, ["start", "coord", "position"])
    print(f"  cols system={sys_col} src={src_col} contig={contig_col} prot={prot_col} coord={coord_col}")
    body = s2_rows[hdr_idx+1:]
    entries = []
    cur = None
    prot_pat = re.compile(r"^[A-Z]{2,4}\d{4,}\.\d+$")
    contig_pat = re.compile(r"^[A-Z]{4}\d{8}$")
    for r in body:
        if not r or all(c is None for c in r): continue
        vals = [None if c is None else str(c).strip() for c in r]
        # detect: is this a new system?
        sysv = vals[sys_col] if sys_col is not None and sys_col < len(vals) else None
        if sysv and sysv.startswith("PD"):
            if cur: entries.append(cur)
            cur = {"pd": sysv, "source": None, "contig": None, "proteins": [], "coord": None}
            if src_col is not None and src_col < len(vals):
                cur["source"] = vals[src_col]
            if contig_col is not None and contig_col < len(vals):
                v = vals[contig_col]
                if v and contig_pat.match(v.split(".")[0]):
                    cur["contig"] = v
            if coord_col is not None and coord_col < len(vals) and vals[coord_col]:
                try:
                    cur["coord"] = int(re.sub(r"[^\d]", "", vals[coord_col]))
                except: pass
        # capture any protein-like accession anywhere in the row
        for v in vals:
            if v and prot_pat.match(v):
                if cur is not None and v not in cur["proteins"]:
                    cur["proteins"].append(v)
            if v and contig_pat.match(v.split(".")[0].split(",")[0]) and cur is not None and cur.get("contig") is None:
                cur["contig"] = v.split(",")[0]
    if cur: entries.append(cur)
    print(f"S2 systems parsed: {len(entries)}")
    total_prot = sum(len(e["proteins"]) for e in entries)
    print(f"S2 total protein components: {total_prot}")
    with open(OUT / "indep_s2_systems.json", "w") as fh:
        json.dump(entries, fh, indent=2)
    print("first 3:", entries[:3])

# ---- Table S4: Gao 2020 comparison ----
s4_rows = None; s4_name = None
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    text_all = " ".join(str(c) for r in rows[:40] for c in r if c is not None).lower()
    if "gao" in text_all and ("identity" in text_all or "percent" in text_all or "cluster" in text_all or "seed" in text_all):
        s4_rows = rows; s4_name = name; break
print(f"\n--- S4 candidate sheet: {s4_name} ---")
if s4_rows:
    for r in s4_rows[:6]:
        print("  ", r)
    # count rows with a numeric identity vs blanks
    hdr_idx = 0
    for i, r in enumerate(s4_rows[:6]):
        if r and any(c and "gao" in str(c).lower() for c in r):
            hdr_idx = i; break
    header = s4_rows[hdr_idx]
    id_col = find_col(header, ["identity", "%", "percent"])
    print(f"S4 header row {hdr_idx}: id_col={id_col}")
    body = s4_rows[hdr_idx+1:]
    with_match, without_match, ids = 0, 0, []
    for r in body:
        if not r or all(c is None for c in r): continue
        row_txt = " ".join(str(c) for c in r if c is not None)
        # need at least a PD protein reference or a component name in first col
        if not any(c for c in r): continue
        if id_col is not None and id_col < len(r):
            v = r[id_col]
            if v is None or (isinstance(v,str) and v.strip() in ("", "-", "N/A", "None", "n/a")):
                without_match += 1
            else:
                try:
                    fv = float(v)
                    with_match += 1
                    ids.append(fv)
                except:
                    without_match += 1
    print(f"S4 components with Gao match: {with_match}")
    print(f"S4 components without Gao match: {without_match}")
    if ids:
        print(f"S4 identities: min={min(ids):.1f} max={max(ids):.1f} n<35={sum(1 for i in ids if i<35)}")
    with open(OUT / "indep_s4_novelty.json", "w") as fh:
        json.dump({"with_match": with_match, "without_match": without_match, "identities": ids}, fh, indent=2)

print("\nDone independent supplement parse.")
