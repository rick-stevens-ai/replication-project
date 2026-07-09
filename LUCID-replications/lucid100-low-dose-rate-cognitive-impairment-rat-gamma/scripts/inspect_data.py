#!/usr/bin/env python3
"""Inspect all supplementary xlsx files to map sheets/columns."""
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parents[1]
ART = REPO / "artifacts"

files = [
    ART/"Data_Sheet_1/Raw data-1/Figure 1-raw data/behavioral tests.xlsx",
    ART/"Data_Sheet_1/Raw data-1/Figure 2-raw data/PETMR-SUVmax.xlsx",
    ART/"Data_Sheet_1/Raw data-1/Figure 3-raw data/SPECTCT-Avg Count.xlsx",
    ART/"Data_Sheet_1/Raw data-1/Figure 4-raw data/Nissl staining.xlsx",
    ART/"Data_Sheet_2/Raw data-2/Figure 5-raw data/SYP.xlsx",
    ART/"Data_Sheet_2/Raw data-2/Figure 6-raw data-1/IF-FJB Iba-CD68 GFAPC3.xlsx",
    ART/"Data_Sheet_3/Raw data-3/Figure 7-raw data/Statistical table of DEGs.xlsx",
    ART/"Data_Sheet_3/Raw data-3/Figure 8-raw data/WB.xlsx",
    ART/"Data_Sheet_3/Raw data-3/Figure 8-raw data/Annotation of KEGG pathway classification for DEGs HDR vs Control.xlsx",
    ART/"Data_Sheet_3/Raw data-3/Figure 8-raw data/Annotation of KEGG pathway classification for DEGs LDR vs Control.xlsx",
]

for f in files:
    print("="*80)
    print("FILE:", f.relative_to(REPO))
    wb = openpyxl.load_workbook(f, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  -- sheet: {sn!r}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
        # print first 6 rows up to col 16
        for r in range(1, min(ws.max_row, 8)+1):
            row = []
            for c in range(1, min(ws.max_column, 16)+1):
                v = ws.cell(r, c).value
                if isinstance(v, float):
                    v = round(v, 4)
                row.append(repr(v))
            print(f"     r{r}: " + " | ".join(row))
