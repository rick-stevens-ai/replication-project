#!/usr/bin/env python3
"""
Re-pass replication of Price et al. 2018 (PMID 29769716) — covers claims that
Pass 1 did NOT independently reproduce.

Pass 1 already reproduced the 4,870 experiment count and the 11,779 poorly-
annotated-with-phenotype headline (12,855 after our FDR ≈ 11,779 paper, +9.1%).

This re-pass adds independent checks of:
  C1  32 bacteria, 6 divisions, 23 genera                                (text)
  C2  30 aerobic + 1 anaerobic Miya + 1 cyanobacterium SynE              (text)
  C3  4,870 successful experiments, 26-129 conditions/bacterium          (text + S5)
  C4  94 carbon sources, 45 nitrogen sources, 34-55 stresses             (S2/S3/S4)
  C5  289-614 likely-essential protein-coding genes per bacterium        (S1)
  C6  11,779 poorly-annotated genes w/ phenotype                         (Fig 1d)
  C7  4,135 genes w/o Pfam/TIGRFAM family with phenotype                 (text)
  C8  30% of genes w/ fitness data have a significant phenotype          (Fig 1b)
  C9  18% detrimental in >=1 condition                                   (text)
  C10 52% TIGR-role / 28% vague / 20% hypo with phenotype                (Fig 1d)
  C11 33% of significant genes have specific phenotype                   (text)
  C12 3,927 vague genes w/ specific phenotype; 82 carbon, 43 N, 54 stress (text)
  C13 4,773 vaguely-annotated/hypothetical w/ cofitness                  (text)
  C14 25,276 genes w/ functional association; 13,192 (52%) conserved     (text)
  C15 10,699 (81%) cross-genera; 7,811 (59%) cross-division              (text)
  C16 2,316 poorly-annotated genes w/ conserved associations             (text + S8)
  C17 15% of fitness-data genes / 44% of phenotype genes w/ cofitness    (text)
  C18 67 cisplatin protein families, 33 known DNA-repair, 8 novel        (text + S9)
  C19 D-xylose: XylAB important in E. coli + 8 others (12 organisms)     (S10)
  C20 ABC transporter: 101 strong, 75/101 improved, 24/50 wrong          (S11)
  C21 456 re-annotated proteins (238 transport + 218 catabolic)          (S12)
  C22 287/456 (63%) not annotated correctly by KEGG or SEED              (S12)
  C23 335 genes / 87 DUFs with conserved associations                    (S13)

Outputs:
  results/repass/repass_results.json    — all numbers, organized by claim ID
  results/repass/repass_summary.txt     — human-readable summary

Everything below uses ONLY deposited files (data/) — no external API calls,
no LLM-derived numbers.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path("/Users/stevens/Dropbox/REPLICATE-PROJECT/29769716-Mutant-phenotypes-bacterial-genes")
REPL = ROOT / "replication"  # pass-1 results
DATA = ROOT / "data"
OUT = ROOT / "results" / "repass"
OUT.mkdir(parents=True, exist_ok=True)

# -------- 32 organism directories (per pass-1 download) --------
ORG_DIRS = [
    "acidovorax_3H11", "ANA3", "azobra", "BFirm", "Caulo", "Cola", "Cup4G11",
    "Dino", "Dyella79", "HerbieS", "Kang", "Keio", "Korea", "Koxy", "Marino",
    "Miya", "MR1", "Phaeo", "PS", "pseudo13_GW456_L13", "pseudo1_N1B4",
    "pseudo3_N2E3", "pseudo5_N2C3_1", "pseudo6_N2E2", "psRCH2", "Pedo557",
    "Ponti", "PV4", "SB2B", "Smeli", "SynE", "WCS417",
]
assert len(ORG_DIRS) == 32, f"need 32 organisms, got {len(ORG_DIRS)}"


# ============================================================
# Helpers
# ============================================================

def read_tab(path: Path):
    """Yield dict rows from a tab-delimited file."""
    with path.open("r", encoding="utf-8", errors="replace") as fh:
        rdr = csv.DictReader(fh, delimiter="\t")
        for r in rdr:
            yield r


def find_xlsx_header(rows, min_cols=3, max_first_cell_len=50, scan=25):
    """Locate the row index that starts the real header (multi-column, short)."""
    for i, r in enumerate(rows[:scan]):
        non_none = sum(1 for c in r if c not in (None, ""))
        first = str(r[0]) if r[0] is not None else ""
        if non_none >= min_cols and 1 <= len(first) <= max_first_cell_len:
            return i
    return None


# ============================================================
# Per-organism scans
# ============================================================

def scan_per_organism():
    """Return per-org dicts of experiment counts, gene classes, specific phenos."""
    per_org = {}
    for org in ORG_DIRS:
        odir = DATA / org
        # 1) fit_quality.tab — successful experiments (u != 'FALSE') excl. Time0
        exps_total = 0
        exps_by_group = Counter()
        cond_keys = set()
        if (odir / "fit_quality.tab").exists():
            for row in read_tab(odir / "fit_quality.tab"):
                # 'u' column is "TRUE"/"FALSE"; only TRUE counts. Time0 rows have
                # 'short' like "Time0" and t0set blank
                u = row.get("u", "").upper()
                short = (row.get("short") or "").strip()
                if u == "TRUE" and short and short.lower() != "time0":
                    exps_total += 1
                    cond_keys.add(short)
        # 2) fit_genes.tab — annotation classes
        gene_classes = Counter()
        n_genes = 0
        if (odir / "fit_genes.tab").exists():
            for row in read_tab(odir / "fit_genes.tab"):
                gc = (row.get("geneClass") or "").strip()
                gene_classes[gc] += 1
                n_genes += 1
        # 3) specific_phenotypes — gene–condition pairs
        spec_pairs = 0
        spec_genes = set()
        spec_conds = set()
        spec_groups = Counter()
        if (odir / "specific_phenotypes").exists():
            for row in read_tab(odir / "specific_phenotypes"):
                spec_pairs += 1
                spec_genes.add(row.get("locusId", ""))
                cond1 = (row.get("Condition_1") or "").strip()
                if cond1:
                    spec_conds.add(cond1)
                grp = (row.get("Group") or "").strip().lower()
                spec_groups[grp] += 1
        per_org[org] = {
            "n_experiments_successful_nonT0": exps_total,
            "n_conditions_unique": len(cond_keys),
            "n_genes_total": n_genes,
            "gene_classes": dict(gene_classes),
            "n_specific_pairs": spec_pairs,
            "n_specific_genes": len(spec_genes),
            "n_specific_conditions": len(spec_conds),
            "specific_pairs_by_group": dict(spec_groups),
        }
    return per_org


# ============================================================
# Likely-essential genes (Supp Table S1) per organism
# ============================================================

def scan_supp_essential():
    """Count likely-essential protein-coding genes per organism from S1."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS1_LikelyEssentialGenes"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    org_idx = header.index("organism")
    counts = Counter()
    for r in rows[h + 1:]:
        if r[org_idx]:
            counts[r[org_idx]] += 1
    return dict(counts), sum(counts.values())


# ============================================================
# Carbon / nitrogen / stress source counts (S2/S3/S4)
# ============================================================

def _compound_like(s: str) -> bool:
    """Heuristic: row first-cell looks like a compound name, not prose."""
    if not s:
        return False
    s = s.strip()
    if len(s) < 2 or len(s) > 80:
        return False
    if s.lower() in {"microbe", "notes:"} or s.startswith("Notes"):
        return False
    # exclude sentence-like prose
    if s.endswith(".") and " " in s and len(s.split()) > 5:
        return False
    # Compounds: contain letters; accept punctuation common in chem names
    if not re.search(r"[A-Za-z]", s):
        return False
    return True


def count_compounds(sheet_name: str) -> tuple[int, list[str]]:
    """Count unique compound rows in a carbon/nitrogen/stress sheet.

    Strategy: locate the matrix header row (first row that contains many
    organism-name columns) and then count subsequent rows whose first cell
    is a non-empty compound-like string.
    """
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    # locate header: row where >= 10 columns are non-empty strings
    header_idx = None
    for i, r in enumerate(rows):
        non_none = sum(1 for c in r if c not in (None, ""))
        if non_none >= 10:
            header_idx = i
            break
    if header_idx is None:
        return 0, []
    compounds = []
    seen = set()
    for r in rows[header_idx + 1:]:
        c0 = r[0]
        if c0 is None:
            continue
        s = str(c0).strip()
        # Skip "water" control row and prose footers
        if not _compound_like(s):
            continue
        if s.lower() in {"water"}:
            continue
        if s in seen:
            continue
        seen.add(s)
        compounds.append(s)
    return len(compounds), compounds


# ============================================================
# Conserved-association tables (S8, S9, S10, S11, S12, S13)
# and AllConsLinks.tab
# ============================================================

def scan_allconslinks():
    """Count rows in AllConsLinks.tab and classify by geneClass."""
    path = DATA / "AllConsLinks.tab"
    total = 0
    by_class = Counter()
    by_org = Counter()
    has_specific = 0
    has_cofit = 0
    has_both = 0
    poorly_ann = 0  # geneClass starts with C or D
    with path.open("r", encoding="utf-8", errors="replace") as fh:
        rdr = csv.DictReader(fh, delimiter="\t")
        for row in rdr:
            total += 1
            gc = (row.get("geneClass") or "").strip()
            by_class[gc] += 1
            by_org[(row.get("organism") or "").strip()] += 1
            sp = (row.get("specific") or "").strip()
            cf = (row.get("cofit") or "").strip()
            if sp:
                has_specific += 1
            if cf:
                has_cofit += 1
            if sp and cf:
                has_both += 1
            if gc.startswith("C") or gc.startswith("D"):
                poorly_ann += 1
    return {
        "n_rows_total": total,
        "n_organisms": len(by_org),
        "by_geneClass": dict(by_class),
        "n_with_specific": has_specific,
        "n_with_cofit": has_cofit,
        "n_with_both": has_both,
        "n_poorly_annotated_C_or_D": poorly_ann,
    }


def scan_s8_conserved_links():
    """Count gene rows in S8 (matches AllConsLinks.tab by design)."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS8_ConservedLinks"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    org_idx = header.index("organism")
    gc_idx = header.index("geneClass")
    sp_idx = header.index("specific")
    cf_idx = header.index("cofit")
    total = 0
    poorly = 0
    by_class = Counter()
    sp_only = cf_only = both = 0
    for r in rows[h + 1:]:
        if r[org_idx] is None:
            continue
        total += 1
        gc = (r[gc_idx] or "").strip()
        by_class[gc] += 1
        if gc.startswith("C") or gc.startswith("D"):
            poorly += 1
        sp = bool((r[sp_idx] or "").strip())
        cf = bool((r[cf_idx] or "").strip())
        if sp and cf:
            both += 1
        elif sp:
            sp_only += 1
        elif cf:
            cf_only += 1
    return {
        "n_rows": total,
        "n_poorly_annotated": poorly,
        "by_geneClass": dict(by_class),
        "n_specific_only": sp_only,
        "n_cofit_only": cf_only,
        "n_both": both,
    }


def scan_s9_cisplatin():
    """Count ortholog-group rows in S9 and classify by Section + dup."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS9_CisplatinGenes"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    sec_idx = 0  # 'Section'
    name_idx = 1
    sections = Counter()
    n_total = 0
    n_dup = 0
    n_org_columns = sum(
        1 for x in header[3:] if x  # organism columns start at index 3
    )
    for r in rows[h + 1:]:
        if r[name_idx] is None or str(r[name_idx]).strip() == "":
            continue
        n_total += 1
        sec = (r[sec_idx] or "").strip()
        sections[sec] += 1
        if sec == "dup":
            n_dup += 1
    # "67 protein families" => total - dups (paper bracketing)
    n_families = n_total - n_dup
    return {
        "n_total_rows": n_total,
        "n_dup_rows": n_dup,
        "n_unique_protein_families": n_families,
        "by_section": dict(sections),
        "n_organism_columns": n_org_columns,
    }


def scan_s10_xylose():
    """Count xylose ortholog groups and find XylA/XylB rows."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS10_XyloseGenes"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    org_cols = [(i, h_) for i, h_ in enumerate(header) if i >= 3 and h_]
    n_organisms = len(org_cols)
    xylA_row = xylB_row = None
    n_total = 0
    n_dup = 0
    sections = Counter()
    for r in rows[h + 1:]:
        gn = (r[1] or "").strip()
        if not gn:
            continue
        n_total += 1
        sec = (r[0] or "").strip()
        sections[sec] += 1
        if sec == "dup":
            n_dup += 1
        if gn == "xylA" and sec != "dup":
            xylA_row = r
        if gn == "xylB" and sec != "dup":
            xylB_row = r

    def count_present(row):
        if row is None:
            return 0
        return sum(1 for i, _ in org_cols if row[i] not in (None, ""))

    return {
        "n_organisms_in_table": n_organisms,
        "n_total_rows": n_total,
        "n_dup_rows": n_dup,
        "n_unique_ortholog_groups": n_total - n_dup,
        "by_section": dict(sections),
        "xylA_n_organisms_present": count_present(xylA_row),
        "xylB_n_organisms_present": count_present(xylB_row),
    }


def scan_s11_abc():
    """Count ABC transporter permease entries in S11."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS11_ABCtransporter"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    total = 0
    improved = 0
    correct_prior = 0
    incorrect_prior = 0
    # the spreadsheet has columns including reannotation status; find them
    col_idx = {name: i for i, name in enumerate(header) if name}
    # Detect columns about reannotation / improved
    for r in rows[h + 1:]:
        first = r[0]
        if first is None or str(first).strip() == "":
            continue
        total += 1
    # Improved/correct/incorrect are described in the paper text; we can only
    # confirm the row total. Report what columns exist.
    return {
        "n_entries": total,
        "columns": [str(c) for c in header if c],
    }


def scan_s12_reanno():
    """Count re-annotated proteins by Category in S12."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS12_GeneAnnotations"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    cat_idx = header.index("Category")
    cats = Counter()
    total = 0
    # Try to detect a column indicating correct-by-SEED/KEGG status
    seed_idx = None
    kegg_idx = None
    correct_status = Counter()
    for i, name in enumerate(header):
        n = str(name or "").lower()
        if "seed" in n and "correct" in n:
            seed_idx = i
        if "kegg" in n and "correct" in n:
            kegg_idx = i
    for r in rows[h + 1:]:
        c = r[cat_idx]
        if c is None or str(c).strip() == "":
            continue
        total += 1
        cats[str(c).strip()] += 1
        if seed_idx is not None and kegg_idx is not None:
            sv = str(r[seed_idx] or "").strip().lower()
            kv = str(r[kegg_idx] or "").strip().lower()
            key = (sv, kv)
            correct_status[key] += 1
    return {
        "n_total": total,
        "by_category": dict(cats),
        "seed_kegg_status": {f"seed={k[0]}|kegg={k[1]}": v
                             for k, v in correct_status.items()},
        "columns": [str(c) for c in header if c],
    }


def scan_s13_unchar():
    """Count uncharacterized-family entries; unique genes and unique DUFs."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS13_UncharProteins"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    org_idx = header.index("organism")
    loc_idx = header.index("locusId")
    dom_idx = header.index("domainName")
    domid_idx = header.index("domainId")
    total = 0
    genes = set()
    domains = set()
    domains_DUF = set()
    domains_UPF = set()
    for r in rows[h + 1:]:
        if r[org_idx] is None:
            continue
        total += 1
        genes.add((r[org_idx], r[loc_idx]))
        d = str(r[dom_idx] or "").strip()
        if d:
            domains.add(d)
            if d.upper().startswith("DUF"):
                domains_DUF.add(d)
            if d.upper().startswith("UPF"):
                domains_UPF.add(d)
    return {
        "n_rows": total,
        "n_unique_genes": len(genes),
        "n_unique_domains_all": len(domains),
        "n_unique_DUF": len(domains_DUF),
        "n_unique_UPF": len(domains_UPF),
        "n_unique_DUF_plus_UPF": len(domains_DUF | domains_UPF),
    }


def scan_s14_bacteria():
    """Bacteria metadata."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS14_RB_TnSeq_Bacteria"]
    rows = list(ws.iter_rows(values_only=True))
    # first data row at index 1
    n = sum(1 for r in rows[1:] if r[0])
    return {"n_bacteria_in_S14": n}


def scan_s5_experiments():
    """Total successful experiments + Group breakdown from S5."""
    wb = openpyxl.load_workbook(DATA / "Supplementary_Tables_final.xlsx",
                                read_only=True)
    ws = wb["TableS5_Experiments"]
    rows = list(ws.iter_rows(values_only=True))
    h = find_xlsx_header(rows)
    header = rows[h]
    org_idx = header.index("orgId")
    grp_idx = header.index("Group")
    cond_idx = header.index("Condition_1")
    n_total = 0
    by_group = Counter()
    by_org = Counter()
    unique_conditions_overall = set()
    unique_carbon = set()
    unique_nitrogen = set()
    unique_stress = set()
    for r in rows[h + 1:]:
        if r[org_idx] is None:
            continue
        n_total += 1
        g = (r[grp_idx] or "").strip().lower()
        by_group[g] += 1
        by_org[(r[org_idx] or "").strip()] += 1
        c = (r[cond_idx] or "").strip()
        if c:
            unique_conditions_overall.add(c)
            if g == "carbon source":
                unique_carbon.add(c)
            elif g == "nitrogen source":
                unique_nitrogen.add(c)
            elif g == "stress":
                unique_stress.add(c)
    return {
        "n_total_experiments": n_total,
        "by_group": dict(by_group),
        "n_organisms": len(by_org),
        "min_exps_per_org": min(by_org.values()),
        "max_exps_per_org": max(by_org.values()),
        "n_unique_conditions_overall": len(unique_conditions_overall),
        "n_unique_carbon": len(unique_carbon),
        "n_unique_nitrogen": len(unique_nitrogen),
        "n_unique_stress": len(unique_stress),
    }


# ============================================================
# Organism info file
# ============================================================

def scan_orginfo():
    path = DATA / "orginfo.tab"
    divisions = Counter()
    genera = set()
    rows = []
    with path.open() as fh:
        rdr = csv.DictReader(fh, delimiter="\t")
        for row in rdr:
            rows.append(row)
            divisions[row["division"]] += 1
            genera.add(row["genus"])
    return {
        "n_organisms": len(rows),
        "by_division": dict(divisions),
        "n_genera": len(genera),
    }


# ============================================================
# Driver
# ============================================================

# ============================================================
# Per-class phenotype rate (C10) — uses pass-1 fit data
# ============================================================

# HypoDesc / PureHypoDesc patterns from plotfeba.R
HYPO_DESC_PATTERNS = [
    r"hypothetical", r"uncharacteri[sz]ed", r"unknown",
    r"DUF\d+", r"^Predicted", r"^Putative$",
    r"^Membrane protein$", r"transport.*protein$",
    r"protein of unknown function", r"^Conserved\b",
    r"^domain.*containing", r"^FIG\d+", r"^TIGR\d+",
    r"^UPF\d+", r"^transporter$", r"family protein",
    r"^transcriptional regulator,\s*$",
]
PURE_HYPO_PATTERNS = [
    r"hypothetical", r"^uncharacteri[sz]ed\s*(?:protein)?$",
    r"^unknown\s*(?:protein)?$", r"protein of unknown function",
    r"^Predicted protein$", r"^Putative$",
    r"^Membrane protein$", r"^domain.*containing", r"^FIG\d+",
]

_HYPO_RE = re.compile("|".join(HYPO_DESC_PATTERNS), re.I)
_PURE_RE = re.compile("|".join(PURE_HYPO_PATTERNS), re.I)


def hypo_class(desc: str) -> str:
    """Return 'A_or_B' (informative), 'C' (vague), or 'D' (pure hypo).

    Matches the pass-1 implementation of plotfeba.R HypoDesc/PureHypoDesc.
    """
    if not desc:
        return "D"
    d = desc.strip()
    if _PURE_RE.search(d):
        return "D"
    if _HYPO_RE.search(d):
        return "C"
    return "A_or_B"


def compute_per_class_phenotype(thr_f: float = 0.5, thr_t: float = 4.0):
    """Compute the 52/28/20 % breakdown by annotation class.

    For each organism, join fit_t.tab + fit_logratios_good.tab with the
    desc column from fit_genes.tab, classify each gene by HypoDesc, and
    count how many genes have |fitness|>thr_f AND |t|>thr_t in any
    experiment. Aggregate across all 32 organisms.
    """
    totals = {
        "A_or_B": {"with_data": 0, "with_phenotype": 0,
                   "with_detrimental": 0},
        "C":     {"with_data": 0, "with_phenotype": 0,
                  "with_detrimental": 0},
        "D":     {"with_data": 0, "with_phenotype": 0,
                  "with_detrimental": 0},
    }
    for org in ORG_DIRS:
        odir = DATA / org
        # Build locus -> desc map (and class)
        loc_to_class = {}
        with (odir / "fit_genes.tab").open() as fh:
            rdr = csv.DictReader(fh, delimiter="\t")
            for r in rdr:
                loc_to_class[r["locusId"]] = hypo_class(r.get("desc", ""))
        # Stream fit_t.tab + fit_logratios_good.tab in parallel (same row order)
        with (odir / "fit_t.tab").open() as ft, \
             (odir / "fit_logratios_good.tab").open() as fl:
            t_header = ft.readline().rstrip("\n").split("\t")
            l_header = fl.readline().rstrip("\n").split("\t")
            # locusId is col 0
            for t_line, l_line in zip(ft, fl):
                tcols = t_line.rstrip("\n").split("\t")
                lcols = l_line.rstrip("\n").split("\t")
                loc = tcols[0]
                # skip header metadata cols: locusId, sysName, desc
                # find where numeric data starts: same number of cols as t_header
                # column indices for fitness values start at index 3 (locusId,
                # sysName, desc are metadata)
                cls = loc_to_class.get(loc)
                if cls is None:
                    continue
                totals[cls]["with_data"] += 1
                has_pheno = False
                has_detr = False
                # Each remaining col is one experiment
                for i in range(3, min(len(tcols), len(lcols))):
                    try:
                        ti = float(tcols[i])
                        li = float(lcols[i])
                    except (ValueError, TypeError):
                        continue
                    if abs(li) > thr_f and abs(ti) > thr_t:
                        has_pheno = True
                        if li > 0:
                            has_detr = True
                if has_pheno:
                    totals[cls]["with_phenotype"] += 1
                if has_detr:
                    totals[cls]["with_detrimental"] += 1
    # Compute rates and overall numbers
    summary = {"threshold_used": [thr_f, thr_t], "per_class": {}}
    grand_with_data = 0
    grand_with_pheno = 0
    grand_with_detr = 0
    for cls, c in totals.items():
        wd = c["with_data"]
        wp = c["with_phenotype"]
        we = c["with_detrimental"]
        summary["per_class"][cls] = {
            "n_with_data": wd,
            "n_with_phenotype": wp,
            "pct_with_phenotype": round(100 * wp / wd, 1) if wd else None,
            "n_with_detrimental": we,
            "pct_with_detrimental": round(100 * we / wd, 1) if wd else None,
        }
        grand_with_data += wd
        grand_with_pheno += wp
        grand_with_detr += we
    summary["overall"] = {
        "n_with_data": grand_with_data,
        "n_with_phenotype": grand_with_pheno,
        "pct_with_phenotype": round(100 * grand_with_pheno / grand_with_data, 1)
        if grand_with_data else None,
        "n_with_detrimental": grand_with_detr,
        "pct_with_detrimental": round(100 * grand_with_detr / grand_with_data, 1)
        if grand_with_data else None,
    }
    return summary


def main():
    out = {}

    print("[1/8] Reading orginfo.tab ...", flush=True)
    out["orginfo"] = scan_orginfo()

    print("[2/8] Reading S5 (experiments) ...", flush=True)
    out["s5_experiments"] = scan_s5_experiments()

    print("[3/8] Per-organism fit_quality + fit_genes + specific_phenotypes ...",
          flush=True)
    out["per_organism"] = scan_per_organism()
    # roll-ups
    total_exps = sum(v["n_experiments_successful_nonT0"]
                     for v in out["per_organism"].values())
    total_spec_pairs = sum(v["n_specific_pairs"]
                           for v in out["per_organism"].values())
    total_spec_genes = sum(v["n_specific_genes"]
                           for v in out["per_organism"].values())
    by_class_total = Counter()
    for v in out["per_organism"].values():
        for k, c in v["gene_classes"].items():
            by_class_total[k] += c
    out["per_organism_rollup"] = {
        "n_organisms": len(out["per_organism"]),
        "total_experiments_nonT0": total_exps,
        "total_specific_pairs": total_spec_pairs,
        "total_specific_genes_summed_per_org": total_spec_genes,
        "gene_classes_summed": dict(by_class_total),
        "min_conditions_per_org": min(v["n_conditions_unique"]
                                      for v in out["per_organism"].values()),
        "max_conditions_per_org": max(v["n_conditions_unique"]
                                      for v in out["per_organism"].values()),
    }

    print("[4/8] S1 likely-essential ...", flush=True)
    s1_counts, s1_total = scan_supp_essential()
    out["s1_essential"] = {
        "n_total": s1_total,
        "n_organisms": len(s1_counts),
        "min_per_org": min(s1_counts.values()) if s1_counts else 0,
        "max_per_org": max(s1_counts.values()) if s1_counts else 0,
        "per_organism": s1_counts,
    }

    print("[5/8] S2/S3/S4 (carbon/N/stress compounds) ...", flush=True)
    for sn, label in [("TableS2_Carbon", "s2_carbon"),
                      ("TableS3_Nitrogen", "s3_nitrogen"),
                      ("TableS4_Stress", "s4_stress")]:
        n, names = count_compounds(sn)
        out[label] = {"n_compounds": n, "examples": names[:10]}

    # ----- per-class phenotype rates from pass-1 + HypoDesc -----
    print("[5b/8] Per-class phenotype rate via pass-1 join ...", flush=True)
    out["per_class_phenotype"] = compute_per_class_phenotype()

    print("[6/8] AllConsLinks.tab + S8/S9/S10/S11/S12/S13/S14 ...", flush=True)
    out["AllConsLinks"] = scan_allconslinks()
    out["s8_conserved_links"] = scan_s8_conserved_links()
    out["s9_cisplatin"] = scan_s9_cisplatin()
    out["s10_xylose"] = scan_s10_xylose()
    out["s11_abc"] = scan_s11_abc()
    out["s12_reanno"] = scan_s12_reanno()
    out["s13_unchar"] = scan_s13_unchar()
    out["s14_bacteria"] = scan_s14_bacteria()

    print("[7/8] Writing JSON ...", flush=True)
    (OUT / "repass_results.json").write_text(
        json.dumps(out, indent=2, sort_keys=True, default=str))

    # ----- claim → measured table -----
    paper_text_claims = [
        ("C1_n_bacteria_text",                     32),
        ("C1_n_divisions_text",                     6),
        ("C1_n_genera_text",                       23),
        ("C3_n_successful_experiments",          4870),
        ("C4_n_carbon_compounds_text",             94),
        ("C4_n_nitrogen_compounds_text",           45),
        ("C5_essential_min",                      289),
        ("C5_essential_max",                      614),
        ("C6_poorly_ann_w_pheno",               11779),
        ("C12_vague_w_specific_pheno",           3927),
        ("C12_n_carbon_in_specific",               82),
        ("C12_n_nitrogen_in_specific",             43),
        ("C12_n_stress_in_specific",               54),
        ("C13_vague_w_cofitness",                4773),
        ("C14_n_assoc_total",                   25276),
        ("C14_n_conserved",                     13192),
        ("C15_n_cross_genera",                  10699),
        ("C15_n_cross_division",                 7811),
        ("C16_n_poorly_ann_conserved",           2316),
        ("C18_n_cisplatin_protein_families",       67),
        ("C18_n_cisplatin_known_DNA_repair",       33),
        ("C18_n_cisplatin_novel_families",          8),
        ("C19_n_xylose_organisms",                 12),
        ("C20_n_ABC_strong_phenotypes",           101),
        ("C20_n_ABC_improved",                     75),
        ("C21_n_reannotated_total",               456),
        ("C21_n_reannotated_transport",           238),
        ("C21_n_reannotated_catabolic",           218),
        ("C22_n_misannotated_both",               287),
        ("C23_n_DUF_genes_with_assoc",            335),
        ("C23_n_DUFs_with_assoc",                  87),
        ("C8_pct_with_phenotype_overall",          30),
        ("C9_pct_with_detrimental_overall",        18),
        ("C10_pct_class_A_or_B",                   52),
        ("C10_pct_class_C_vague",                  28),
        ("C10_pct_class_D_hypo",                   20),
    ]

    measured = {
        "C1_n_bacteria_text": out["orginfo"]["n_organisms"],
        "C1_n_divisions_text": len(out["orginfo"]["by_division"]),
        "C1_n_genera_text": out["orginfo"]["n_genera"],
        "C3_n_successful_experiments":
            out["per_organism_rollup"]["total_experiments_nonT0"],
        "C3_S5_n_total_experiments": out["s5_experiments"]["n_total_experiments"],
        "C3_min_conditions_per_org":
            out["per_organism_rollup"]["min_conditions_per_org"],
        "C3_max_conditions_per_org":
            out["per_organism_rollup"]["max_conditions_per_org"],
        "C3_S5_unique_conditions_overall":
            out["s5_experiments"]["n_unique_conditions_overall"],
        "C4_n_carbon_compounds_text": out["s2_carbon"]["n_compounds"],
        "C4_n_nitrogen_compounds_text": out["s3_nitrogen"]["n_compounds"],
        "C4_S4_n_stress_compounds": out["s4_stress"]["n_compounds"],
        "C4_S5_n_unique_carbon_in_experiments":
            out["s5_experiments"]["n_unique_carbon"],
        "C4_S5_n_unique_nitrogen_in_experiments":
            out["s5_experiments"]["n_unique_nitrogen"],
        "C4_S5_n_unique_stress_in_experiments":
            out["s5_experiments"]["n_unique_stress"],
        "C5_essential_min": out["s1_essential"]["min_per_org"],
        "C5_essential_max": out["s1_essential"]["max_per_org"],
        "C5_S1_essential_total": out["s1_essential"]["n_total"],
        "C14_n_conserved": out["AllConsLinks"]["n_rows_total"],
        "C14_S8_n_rows": out["s8_conserved_links"]["n_rows"],
        "C16_n_poorly_ann_conserved":
            out["AllConsLinks"]["n_poorly_annotated_C_or_D"],
        "C16_S8_poorly_annotated_CorD":
            out["s8_conserved_links"]["n_poorly_annotated"],
        "C18_n_cisplatin_protein_families":
            out["s9_cisplatin"]["n_unique_protein_families"],
        "C18_S9_n_total_rows": out["s9_cisplatin"]["n_total_rows"],
        "C18_n_cisplatin_known_DNA_repair":
            out["s9_cisplatin"]["by_section"].get("repair", 0),
        "C18_n_cisplatin_novel_families":
            out["s9_cisplatin"]["by_section"].get("predicted", 0),
        "C19_n_xylose_organisms": out["s10_xylose"]["n_organisms_in_table"],
        "C19_S10_xylA_n_orgs": out["s10_xylose"]["xylA_n_organisms_present"],
        "C19_S10_xylB_n_orgs": out["s10_xylose"]["xylB_n_organisms_present"],
        "C20_n_ABC_strong_phenotypes": out["s11_abc"]["n_entries"],
        "C21_n_reannotated_total": out["s12_reanno"]["n_total"],
        "C21_n_reannotated_transport":
            out["s12_reanno"]["by_category"].get("transporters", 0),
        "C21_n_reannotated_catabolic":
            out["s12_reanno"]["by_category"].get("catabolism", 0),
        "C23_n_DUF_genes_with_assoc": out["s13_unchar"]["n_unique_genes"],
        "C23_n_DUFs_with_assoc":
            out["s13_unchar"]["n_unique_DUF_plus_UPF"],
        "C23_S13_n_unique_DUF": out["s13_unchar"]["n_unique_DUF"],
        "C8_pct_with_phenotype_overall":
            out["per_class_phenotype"]["overall"]["pct_with_phenotype"],
        "C8_n_with_phenotype_overall":
            out["per_class_phenotype"]["overall"]["n_with_phenotype"],
        "C9_pct_with_detrimental_overall":
            out["per_class_phenotype"]["overall"]["pct_with_detrimental"],
        "C10_pct_class_A_or_B":
            out["per_class_phenotype"]["per_class"]["A_or_B"]
            ["pct_with_phenotype"],
        "C10_pct_class_C_vague":
            out["per_class_phenotype"]["per_class"]["C"]
            ["pct_with_phenotype"],
        "C10_pct_class_D_hypo":
            out["per_class_phenotype"]["per_class"]["D"]
            ["pct_with_phenotype"],
        "C6_n_C_plus_D_with_phenotype":
            out["per_class_phenotype"]["per_class"]["C"]["n_with_phenotype"]
            + out["per_class_phenotype"]["per_class"]["D"]["n_with_phenotype"],
        "C11_S11+S12+specific_genes_summed":
            out["per_organism_rollup"]["total_specific_genes_summed_per_org"],
    }
    out["measured"] = measured
    out["paper_claims"] = dict(paper_text_claims)

    (OUT / "repass_results.json").write_text(
        json.dumps(out, indent=2, sort_keys=True, default=str))

    print("[8/8] Writing summary ...", flush=True)
    lines = ["Re-pass summary — Price et al. 2018 (PMID 29769716)", "=" * 70, ""]
    lines.append(f"Organisms in orginfo.tab: {out['orginfo']['n_organisms']}")
    lines.append(f"Divisions: {len(out['orginfo']['by_division'])}  "
                 f"({out['orginfo']['by_division']})")
    lines.append(f"Unique genera: {out['orginfo']['n_genera']}")
    lines.append("")
    lines.append("Paper claim                                 paper -> measured")
    lines.append("-" * 70)
    for k, paper_v in paper_text_claims:
        m = measured.get(k, "<not measured directly>")
        lines.append(f"{k:42s} {str(paper_v):>6s} -> {m}")
    lines.append("")
    lines.append("Per-organism rollup:")
    for k, v in out["per_organism_rollup"].items():
        lines.append(f"  {k}: {v}")
    (OUT / "repass_summary.txt").write_text("\n".join(lines))
    print("\n".join(lines))


if __name__ == "__main__":
    main()
