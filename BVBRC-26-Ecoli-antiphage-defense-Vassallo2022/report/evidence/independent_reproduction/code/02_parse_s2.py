#!/usr/bin/env python3
"""Direct parse of Table S2 (systems with source strain, contig, protein accessions)."""
import json, re
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(__file__).resolve().parents[3].parent / "work" / "SupplementaryTables.xlsx"
OUT = Path(__file__).resolve().parents[1] / "data"

wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb["Table S2"]
rows = list(ws.iter_rows(values_only=True))
print(f"Table S2 rows: {len(rows)}")
for i, r in enumerate(rows[:8]):
    print(i, r)

# Print first ~15 rows fully
print("\nFirst 20 non-empty rows:")
n = 0
for i, r in enumerate(rows):
    if r and any(c is not None for c in r):
        print(i, r)
        n += 1
        if n >= 25: break
